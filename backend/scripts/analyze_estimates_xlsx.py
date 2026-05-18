"""
One-shot analyzer for the school's `הערכות לשנת הלימודים` Excel.

Run: `python backend/scripts/analyze_estimates_xlsx.py path/to/file.xlsx`

The Excel has ~30 sheets — most are "subject sheets" with a fixed header
(כיתה / סוג כיתה / שם המורה המלמד / שעות הוראה / שעות גמול בגרות / סמל שאלון בגרות /
הערות) but with significant variations:

- Some sheets pool multiple classes for ability-level groups
  ("יא 1,2,5,6,7,9" + rows for `3 יח"ל / 4 יח"ל / 5 יח"ל`).
- חינוך גופני splits into מורה לבנים / מורה לבנות per class.
- אנגלית has a side-table summarising per-teacher hours by section.
- תפקידים / השכלה don't follow the per-class layout at all.

This script does not write anything — it prints a structured summary so we
can verify the parser against the real file before we touch the DB.
"""
from __future__ import annotations

import json
import re
import sys
import warnings
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

warnings.filterwarnings('ignore')


GRADE_LETTERS = ['ז', 'ח', 'ט', 'י', 'יא', 'יב']
GRADE_LEVEL = {'ז': 7, 'ח': 8, 'ט': 9, 'י': 10, 'יא': 11, 'יב': 12}
# Some cells use "יו"ד" instead of "י" for grade 10.
GRADE_ALIASES = {
    'יו"ד': 'י', 'יוד': 'י', "י'": 'י',
    'י"א': 'יא', "יא'": 'יא',
    'י"ב': 'יב', "יב'": 'יב',
}
ROLE_SHEETS = {'תפקידים'}
COURSE_CATALOG_SHEETS = {'השכלה'}  # course list, not per-class
PE_SHEETS = {'חינוך גופני'}


# ─────────────────────────────────────────────────────────────────────────────
# Cell parsing helpers
# ─────────────────────────────────────────────────────────────────────────────

CLASS_RE = re.compile(r'(יב|יא|י|ט|ח|ז)\s*[\'"]?\s*(\d+)')
CLASS_LIST_RE = re.compile(r'(יב|יא|י|ט|ח|ז)\s*[\'"]?\s*([\d,\s]+)')


def normalize_str(s):
    if s is None:
        return ''
    return str(s).strip().replace('"', '"').replace("'", "'")


def parse_class_token(text):
    """Parse a single class name like 'ז1' / 'יא 3' / 'יו"ד 5' → ('יא', 3) or None."""
    text = normalize_str(text)
    if not text:
        return None
    for alias, canonical in GRADE_ALIASES.items():
        text = text.replace(alias, canonical)
    m = CLASS_RE.match(text)
    if not m:
        return None
    return m.group(1), int(m.group(2))


def parse_class_list(text):
    """Parse a cell that lists multiple classes for a pool, e.g.
    'יא 1,2,5,6,7,9' → [('יא', 1), ('יא', 2), …]. Returns [] if none found."""
    text = normalize_str(text)
    if not text:
        return []
    for alias, canonical in GRADE_ALIASES.items():
        text = text.replace(alias, canonical)
    # Find a grade prefix once and parse all numbers in that token.
    m = CLASS_LIST_RE.search(text)
    if not m:
        return []
    grade = m.group(1)
    nums = re.findall(r'\d+', m.group(2))
    return [(grade, int(n)) for n in nums]


def parse_hours(value):
    """'3+1' → Decimal('4'), '2.5' → Decimal('2.5'), None → None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = normalize_str(value)
    if not s:
        return None
    # Strip non-numeric prefixes like '(+2 תוכנה?) 6' → '6'
    matches = re.findall(r'\d+(?:\.\d+)?', s)
    if not matches:
        return None
    try:
        total = Decimal('0')
        for m in matches:
            total += Decimal(m)
        # '+' style: sum all (e.g., 3+1=4, 4+4=8). For a single number this is fine.
        return total
    except InvalidOperation:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Sheet classification & header detection
# ─────────────────────────────────────────────────────────────────────────────

HEADER_KEYS = [
    # Order matters: most specific labels first. Many keys share substrings
    # (e.g., "סוג כיתה" contains "כיתה"), so we must check the more specific
    # ones before the generic ones, and prefer exact matches over substrings.
    ('class_type', ['סוג כיתה']),
    ('boys_teacher', ['מורה לבנים']),
    ('girls_teacher', ['מורה לבנות']),
    ('boys_count', ['מספרי בנים', 'מספר בנים']),
    ('girls_count', ['מספרי בנות', 'מספר בנות']),
    ('student_count', ["מס' תלמידים", 'מספר תלמידים']),
    ('teacher', ['שם המורה המלמד', 'שם מורה', 'שם המורה', 'מורה']),
    ('bagrut_hours', ['שעות גמול בגרות', 'גמול בגרות']),
    ('bagrut_code', ['סמל שאלון בגרות', 'סמל שאלון']),
    ('hours', ['שעות הוראה', 'שעות']),
    ('notes', ['הערות']),
    ('class', ['כיתה']),
]


def _match_header_key(text):
    """Return the most specific HEADER_KEYS entry matching `text`, or None.

    Preference order:
      1. exact equality with any candidate;
      2. otherwise, the first key whose candidate appears as a substring.
    """
    if not text:
        return None
    for key, candidates in HEADER_KEYS:
        if any(c == text for c in candidates):
            return key
    for key, candidates in HEADER_KEYS:
        if any(c in text for c in candidates):
            return key
    return None


def find_header_row(ws):
    for r in range(1, min(8, ws.max_row + 1)):
        for cell in ws[r]:
            if normalize_str(cell.value) == 'כיתה':
                return r
    return None


def detect_columns(ws, header_row):
    """Map header text → column index (1-based). First-occurrence-wins per key,
    so a teacher-summary side-table (e.g., the right-hand block in אנגלית)
    won't clobber the primary assignment columns on the left."""
    mapping = {}
    for cell in ws[header_row]:
        text = normalize_str(cell.value)
        key = _match_header_key(text)
        if key is None:
            continue
        mapping.setdefault(key, cell.column)
    return mapping


def detect_subject_name(ws):
    """Subject name lives in row 1, usually in cell B1 or C1."""
    header_strings = set()
    for _, candidates in HEADER_KEYS:
        header_strings.update(candidates)
    for r in range(1, 4):
        for cell in ws[r]:
            text = normalize_str(cell.value)
            if text and text != 'כיתה' and text not in header_strings:
                # Skip purely numeric junk
                if re.fullmatch(r'\d+(?:\.\d+)?', text):
                    continue
                return text
    return ws.title


# ─────────────────────────────────────────────────────────────────────────────
# Parsing rows from a "standard" subject sheet
# ─────────────────────────────────────────────────────────────────────────────

def parse_subject_sheet(ws):
    """Return a list of records, each describing one row in the sheet that
    looks like an assignment / group / class spec.

    Records are intentionally lossy-tolerant — we keep raw values too so the
    importer can decide what to do per record.
    """
    header_row = find_header_row(ws)
    if not header_row:
        return {
            'kind': 'unknown',
            'subject': ws.title,
            'records': [],
            'reason': 'no header row found',
        }

    cols = detect_columns(ws, header_row)
    subject = detect_subject_name(ws)

    is_pe = ws.title in PE_SHEETS or ('boys_teacher' in cols and 'girls_teacher' in cols)

    records = []
    last_class_pool = None  # carry forward the most recent class token for "group" rows
    last_grade = None  # carry forward grade if rows are within a grade-block

    for r in range(header_row + 1, ws.max_row + 1):
        row = ws[r]

        def cv(key):
            col = cols.get(key)
            return row[col - 1].value if col else None

        class_raw = normalize_str(cv('class'))
        class_type_raw = normalize_str(cv('class_type'))
        hours_raw = cv('hours')

        # Skip totally empty rows.
        if not any(normalize_str(c.value) for c in row):
            continue

        # Class column might be a single class, a list, a "הקבצה" marker, or empty.
        class_tokens = []
        pool_label = ''
        if class_raw:
            single = parse_class_token(class_raw)
            multi = parse_class_list(class_raw)
            if single:
                class_tokens = [single]
                last_grade = single[0]
                last_class_pool = [single]
            elif multi:
                class_tokens = multi
                last_grade = multi[0][0]
                last_class_pool = multi
            else:
                pool_label = class_raw

        # Rows with no class token but with class_type or hours are group/level rows
        # that belong to the most recent pool.
        if not class_tokens and last_class_pool and (class_type_raw or hours_raw):
            class_tokens = list(last_class_pool)

        teacher = normalize_str(cv('teacher'))
        boys_teacher = normalize_str(cv('boys_teacher')) if is_pe else ''
        girls_teacher = normalize_str(cv('girls_teacher')) if is_pe else ''
        hours = parse_hours(hours_raw)
        bagrut_hours = parse_hours(cv('bagrut_hours'))
        bagrut_code = normalize_str(cv('bagrut_code'))
        notes = normalize_str(cv('notes'))
        student_count = parse_hours(cv('student_count'))

        # Some sheets have totals rows like "חטע 194 29.75" — skip them.
        is_total = class_raw and class_raw.replace('"', '').strip() in {'חטע', 'חטב', 'חט"ע', 'חט"ב'}

        if not class_tokens and not (teacher or boys_teacher or girls_teacher) and not pool_label:
            continue

        records.append({
            'row': r,
            'classes': class_tokens,
            'pool_label': pool_label,
            'class_type_raw': class_type_raw,
            'teacher': teacher,
            'boys_teacher': boys_teacher,
            'girls_teacher': girls_teacher,
            'hours': str(hours) if hours is not None else None,
            'bagrut_hours': str(bagrut_hours) if bagrut_hours is not None else None,
            'bagrut_code': bagrut_code,
            'notes': notes,
            'student_count': str(student_count) if student_count is not None else None,
            'is_total_row': is_total,
            'is_pe': is_pe,
        })

    return {
        'kind': 'pe' if is_pe else 'subject',
        'subject': subject,
        'header_row': header_row,
        'columns': {k: openpyxl.utils.get_column_letter(v) for k, v in cols.items()},
        'records': records,
    }


def parse_role_sheet(ws):
    """תפקידים sheet. Columns: role-name | context | description | teacher | hours | stipend | required-to-teach."""
    records = []
    header_row = 1
    for r in range(header_row + 1, ws.max_row + 1):
        row = [normalize_str(c.value) for c in ws[r]]
        if not any(row):
            continue
        role = row[0] if len(row) > 0 else ''
        context = row[1] if len(row) > 1 else ''
        desc = row[2] if len(row) > 2 else ''
        teacher = row[3] if len(row) > 3 else ''
        hours = parse_hours(row[4]) if len(row) > 4 else None
        stipend = parse_hours(row[5]) if len(row) > 5 else None
        must_teach = parse_hours(row[6]) if len(row) > 6 else None
        if not role and not teacher:
            continue
        records.append({
            'role': role,
            'context': context,
            'description': desc,
            'teacher': teacher,
            'hours': str(hours) if hours is not None else None,
            'stipend': str(stipend) if stipend is not None else None,
            'must_teach_hours': str(must_teach) if must_teach is not None else None,
        })
    return {'kind': 'roles', 'records': records}


def parse_course_catalog_sheet(ws):
    """השכלה — list of elective courses. Layout is loose; just dump rows."""
    records = []
    for r in range(1, ws.max_row + 1):
        row = [normalize_str(c.value) for c in ws[r]]
        if not any(row):
            continue
        records.append(row[:6])
    return {'kind': 'course_catalog', 'records': records}


def parse_sheet(ws):
    if ws.title in ROLE_SHEETS:
        return parse_role_sheet(ws)
    if ws.title in COURSE_CATALOG_SHEETS:
        return parse_course_catalog_sheet(ws)
    return parse_subject_sheet(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Top-level summary
# ─────────────────────────────────────────────────────────────────────────────

def analyze(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    teachers = Counter()  # teacher → total hours seen (any sheet)
    classes_per_grade = defaultdict(set)
    class_types = Counter()
    bagrut_codes = Counter()
    subjects = []
    role_records = []
    catalog_records = []

    per_sheet = {}

    for name in wb.sheetnames:
        ws = wb[name]
        result = parse_sheet(ws)
        per_sheet[name] = result

        if result['kind'] == 'roles':
            for rec in result['records']:
                if rec['teacher']:
                    h = Decimal(rec['hours']) if rec['hours'] else Decimal('0')
                    teachers[rec['teacher']] += h
            role_records.extend(result['records'])
            continue
        if result['kind'] == 'course_catalog':
            catalog_records.extend(result['records'])
            continue

        subject = result['subject']
        subjects.append(subject)
        for rec in result['records']:
            if rec.get('is_total_row'):
                continue
            for grade, num in rec['classes']:
                classes_per_grade[grade].add(num)
            ct = rec['class_type_raw']
            if ct:
                class_types[ct] += 1
            bc = rec['bagrut_code']
            if bc:
                bagrut_codes[bc] += 1
            hrs = Decimal(rec['hours']) if rec['hours'] else Decimal('0')
            for teacher_field in ('teacher', 'boys_teacher', 'girls_teacher'):
                t = rec[teacher_field]
                if t:
                    teachers[t] += hrs

    summary = {
        'subjects': subjects,
        'subject_count': len(subjects),
        'classes_per_grade': {g: sorted(classes_per_grade[g]) for g in GRADE_LETTERS if classes_per_grade.get(g)},
        'class_type_frequencies': dict(class_types.most_common()),
        'bagrut_codes': dict(bagrut_codes.most_common()),
        'teacher_count': len(teachers),
        'teacher_top20': [
            {'name': t, 'total_hours_seen': str(h)} for t, h in teachers.most_common(20)
        ],
        'role_records_total': len(role_records),
        'catalog_records_total': len(catalog_records),
        'sheets': {
            name: {
                'kind': res['kind'],
                'record_count': len(res.get('records', [])),
                'subject': res.get('subject'),
                'columns': res.get('columns'),
            }
            for name, res in per_sheet.items()
        },
    }
    return summary, per_sheet


def main():
    if len(sys.argv) < 2:
        print('usage: analyze_estimates_xlsx.py <file.xlsx> [--full]')
        sys.exit(2)
    path = sys.argv[1]
    full = '--full' in sys.argv[2:]
    summary, per_sheet = analyze(path)

    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))

    if full:
        print('\n--- PER-SHEET DETAIL ---')
        for name, res in per_sheet.items():
            print(f'\n### {name} ({res["kind"]}, {len(res.get("records", []))} records)')
            for rec in res.get('records', [])[:80]:
                print(json.dumps(rec, ensure_ascii=False, default=str))


if __name__ == '__main__':
    main()
