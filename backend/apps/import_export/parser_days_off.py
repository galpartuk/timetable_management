"""
Excel parser for teacher days off.

Expected format:
- Sheet with two columns: teacher name + day-off marker
- Either a single "day" column with the day name/letter,
  or one column per day (Sun..Thu) with any non-empty cell marking the day off.

Hebrew day letters and names are both accepted: א/ראשון, ב/שני, ג/שלישי, ד/רביעי, ה/חמישי.
"""
from openpyxl import load_workbook

from apps.subjects.models import Teacher


DAY_LOOKUP = {
    'א': 1, 'ראשון': 1, 'sunday': 1, 'sun': 1,
    'ב': 2, 'שני': 2, 'monday': 2, 'mon': 2,
    'ג': 3, 'שלישי': 3, 'tuesday': 3, 'tue': 3,
    'ד': 4, 'רביעי': 4, 'wednesday': 4, 'wed': 4,
    'ה': 5, 'חמישי': 5, 'thursday': 5, 'thu': 5,
}


def _parse_day(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        d = int(value)
        return d if 1 <= d <= 5 else None
    s = str(value).strip().lower()
    if not s:
        return None
    return DAY_LOOKUP.get(s) or DAY_LOOKUP.get(s[:1])


def _detect_layout(ws):
    """Returns ('two_col', name_col, day_col) or ('wide', name_col, {col: day})."""
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        if any(c.value for c in row):
            header_row = row
            break
    if header_row is None:
        return None

    name_col = None
    day_col = None
    wide_cols = {}
    for cell in header_row:
        if cell.value is None:
            continue
        text = str(cell.value).strip()
        lower = text.lower()
        if any(kw in text for kw in ('מורה', 'שם')) or 'teacher' in lower or 'name' in lower:
            name_col = cell.column
            continue
        if 'יום' in text or 'day' in lower:
            day_col = cell.column
            continue
        d = _parse_day(text)
        if d:
            wide_cols[cell.column] = d

    if name_col and day_col:
        return ('two_col', name_col, day_col, header_row[0].row)
    if name_col and wide_cols:
        return ('wide', name_col, wide_cols, header_row[0].row)
    return None


def _find_teacher(school, name):
    name = name.strip()
    if not name:
        return None
    # Match against full name first, then first_name only (importer creates teachers with first_name=fullname)
    qs = Teacher.objects.filter(school=school)
    for t in qs:
        if str(t).strip() == name or t.first_name.strip() == name:
            return t
    return None


def parse_days_off_excel(file, school, import_log):
    wb = load_workbook(file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    layout = _detect_layout(ws)
    if layout is None:
        raise ValueError('לא נמצאה כותרת מתאימה (שם מורה / יום)')

    updated = 0
    errors: list[str] = []

    if layout[0] == 'two_col':
        _, name_col, day_col, header_row_idx = layout
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
            name_val = row[name_col - 1].value
            day_val = row[day_col - 1].value
            if not name_val:
                continue
            day = _parse_day(day_val)
            if day is None:
                errors.append(f'{name_val}: יום לא חוקי "{day_val}"')
                continue
            teacher = _find_teacher(school, str(name_val))
            if teacher is None:
                errors.append(f'מורה לא נמצא: {name_val}')
                continue
            teacher.day_off = day
            teacher.save(update_fields=['day_off'])
            updated += 1
    else:
        _, name_col, wide_cols, header_row_idx = layout
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
            name_val = row[name_col - 1].value
            if not name_val:
                continue
            chosen = None
            for col, day in wide_cols.items():
                if row[col - 1].value not in (None, ''):
                    chosen = day
                    break
            if chosen is None:
                continue
            teacher = _find_teacher(school, str(name_val))
            if teacher is None:
                errors.append(f'מורה לא נמצא: {name_val}')
                continue
            teacher.day_off = chosen
            teacher.save(update_fields=['day_off'])
            updated += 1

    import_log.teachers_imported = updated
    import_log.errors = errors
    import_log.save()

    return {'teachers_updated': updated, 'errors': errors}
