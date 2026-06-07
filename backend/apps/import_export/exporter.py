"""Excel exporter — builds a multi-sheet workbook from a `spec` dict.

The spec is a flat dict like:
    {
        'timetable_id': 1,
        'school_id': 1,
        'sheets': [
            'timetable_by_class', 'timetable_by_teacher', 'timetable_flat',
            'teachers', 'subjects', 'classes', 'time_slots', 'rooms',
            'assignments', 'constraints', 'conflicts',
            'import_logs', 'audit_logins', 'audit_activities', 'users',
        ],
    }

Each sheet type maps to a `_build_*` function below. New sheet = add a
function and register it in SHEET_BUILDERS. The view layer enforces
permission checks before audit/users sheets are produced.
"""
from __future__ import annotations

from collections import defaultdict
from typing import Any, Callable, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ── styling helpers ──────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='4F46E5')
SUBHEADER_FILL = PatternFill('solid', fgColor='F1F3F7')
GRID_FONT = Font(size=10)
WRAP = Alignment(wrap_text=True, vertical='center', horizontal='center')


def _style_header(ws: Worksheet, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _autosize(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        widths = [len(str(c.value)) if c.value is not None else 0 for c in col]
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(min_width, max(widths) + 2), max_width,
        )


def _write_table(ws: Worksheet, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = 'A2'
    for row in rows:
        ws.append(row)
    _autosize(ws)


# ── individual sheet builders ────────────────────────────────────────────

def _build_teachers(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.subjects.models import Teacher
    school_id = spec.get('school_id') or 1
    qs = Teacher.objects.filter(school_id=school_id).order_by('last_name', 'first_name')
    rows = [[
        t.id, t.first_name, t.last_name or '', t.email or '', t.phone or '',
        t.max_weekly_hours, t.get_day_off_display() if t.day_off else '',
    ] for t in qs]
    ws = wb.create_sheet('מורים')
    _write_table(ws, [
        'ID', 'שם פרטי', 'שם משפחה', 'אימייל', 'טלפון', 'מקס שעות שבועי', 'יום חופש',
    ], rows)


def _build_subjects(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.subjects.models import Subject
    school_id = spec.get('school_id') or 1
    qs = Subject.objects.filter(school_id=school_id).order_by('name_he')
    rows = [[s.id, s.name_he, s.name_en, s.color, 'כן' if s.requires_consecutive else ''] for s in qs]
    ws = wb.create_sheet('מקצועות')
    _write_table(ws, ['ID', 'שם בעברית', 'שם באנגלית', 'צבע', 'דורש שיעורים צמודים'], rows)
    # Tint each row's color column with the actual subject color
    for r_idx, s in enumerate(qs, start=2):
        try:
            ws.cell(row=r_idx, column=4).fill = PatternFill('solid', fgColor=s.color.lstrip('#'))
        except Exception:
            pass


def _build_classes(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.school.models import SchoolClass
    school_id = spec.get('school_id') or 1
    qs = SchoolClass.objects.filter(grade__school_id=school_id).select_related('grade')
    rows = [[
        c.id, c.grade.name, c.number, c.display_name, c.get_class_type_display(), c.student_count,
    ] for c in qs]
    ws = wb.create_sheet('כיתות')
    _write_table(ws, ['ID', 'שכבה', 'מספר', 'שם תצוגה', 'סוג', 'מספר תלמידים'], rows)


def _build_time_slots(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.school.models import TimeSlot
    school_id = spec.get('school_id') or 1
    qs = TimeSlot.objects.filter(school_id=school_id).order_by('day', 'period')
    rows = [[
        ts.id, ts.day, ts.get_day_display(), ts.period,
        ts.start_time.strftime('%H:%M'), ts.end_time.strftime('%H:%M'),
    ] for ts in qs]
    ws = wb.create_sheet('משבצות זמן')
    _write_table(ws, ['ID', 'יום (מספר)', 'יום (שם)', 'שיעור', 'התחלה', 'סיום'], rows)


def _build_rooms(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.school.models import Room
    school_id = spec.get('school_id') or 1
    qs = Room.objects.filter(school_id=school_id)
    rows = [[r.id, r.name, r.capacity, r.room_type] for r in qs]
    ws = wb.create_sheet('חדרים')
    _write_table(ws, ['ID', 'שם', 'קיבולת', 'סוג'], rows)


def _build_assignments(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.subjects.models import TeachingAssignment
    school_id = spec.get('school_id') or 1
    qs = (
        TeachingAssignment.objects
        .filter(school_class__grade__school_id=school_id)
        .select_related('teacher', 'subject', 'school_class', 'school_class__grade')
    )
    rows = [[
        a.id,
        f'{a.teacher.first_name} {a.teacher.last_name or ""}'.strip(),
        a.subject.name_he,
        a.school_class.display_name,
        float(a.weekly_hours),
        float(a.bagrut_bonus_hours),
        a.bagrut_exam_code or '',
        a.notes or '',
    ] for a in qs]
    ws = wb.create_sheet('שיבוצי הוראה')
    _write_table(ws, [
        'ID', 'מורה', 'מקצוע', 'כיתה', 'שעות שבועיות',
        'גמול בגרות', 'סמל שאלון', 'הערות',
    ], rows)


def _build_teacher_summary(wb: Workbook, spec: Dict[str, Any]) -> None:
    """What the system understood per teacher: one row per (teacher, subject,
    class) with hours, plus the teacher's department. Pooled classes are
    expanded so every class a teacher reaches shows up. This is the human
    check that the import read the workbook correctly."""
    from apps.subjects.models import Teacher, TeachingAssignment
    school_id = spec.get('school_id') or 1
    teachers = {
        t.id: t for t in Teacher.objects.filter(school_id=school_id).prefetch_related('tags')
    }
    qs = (
        TeachingAssignment.objects
        .filter(school_class__grade__school_id=school_id)
        .select_related('subject', 'school_class', 'school_class__grade')
        .prefetch_related('additional_classes__grade')
    )
    rows: List[List[Any]] = []
    for a in qs:
        t = teachers.get(a.teacher_id)
        if not t:
            continue
        dept = ', '.join(tag.name for tag in t.tags.all() if tag.kind == 'department')
        name = f'{t.first_name} {t.last_name or ""}'.strip()
        for cls in [a.school_class, *a.additional_classes.all()]:
            rows.append([
                name, dept, a.subject.name_he, cls.display_name,
                float(a.weekly_hours), 'כן' if a.is_active else 'לא',
            ])
    rows.sort(key=lambda r: (r[0], r[2], r[3]))
    ws = wb.create_sheet('סיכום מורים')
    _write_table(ws, ['מורה', 'מחלקה', 'מקצוע', 'כיתה', 'שעות שבועיות', 'פעיל'], rows)


def _build_diagnostics(wb: Workbook, spec: Dict[str, Any]) -> None:
    """Pre-flight feasibility: per-class/teacher load vs available slots, run
    through the same checker the solver uses. Surfaces overloaded classes
    (e.g. a class needing more lessons than the week has slots) in the file so
    the user can see and fix them before building."""
    from apps.school.models import School
    from solver.feasibility import analyze
    school_id = spec.get('school_id') or 1
    ws = wb.create_sheet('אבחון')
    school = School.objects.filter(id=school_id).first()
    if not school:
        _write_table(ws, ['חומרה', 'קוד', 'יעד', 'הודעה'], [])
        return
    report = analyze(school)
    rows: List[List[Any]] = []
    for label, issues in (
        ('חוסם', report.blockers), ('אזהרה', report.warnings), ('מידע', report.info),
    ):
        for i in issues:
            rows.append([label, i.code, i.target, i.message])
    _write_table(ws, ['חומרה', 'קוד', 'יעד', 'הודעה'], rows)


def _build_constraints(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.scheduling.models import Constraint
    school_id = spec.get('school_id') or 1
    qs = Constraint.objects.filter(school_id=school_id).select_related('teacher', 'subject', 'school_class')
    rows = [[
        c.id, c.name, c.get_constraint_type_display(), c.get_priority_display(),
        'כן' if c.is_active else 'לא',
        str(c.teacher) if c.teacher else '',
        c.subject.name_he if c.subject else '',
        c.school_class.display_name if c.school_class else '',
        str(c.parameters or {}),
    ] for c in qs]
    ws = wb.create_sheet('אילוצים')
    _write_table(ws, [
        'ID', 'שם', 'סוג', 'עדיפות', 'פעיל', 'מורה', 'מקצוע', 'כיתה', 'פרמטרים',
    ], rows)


# ── timetable views ──────────────────────────────────────────────────────

DAY_NAMES = {1: 'ראשון', 2: 'שני', 3: 'שלישי', 4: 'רביעי', 5: 'חמישי'}


def _grid_for_entries(ws: Worksheet, title: str, entries, mode: str) -> None:
    """Render a day×period grid for a single class or teacher.

    `mode='class'` → cell shows subject + teacher.
    `mode='teacher'` → cell shows subject + class.
    """
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    headers = ['שיעור'] + [DAY_NAMES[d] for d in (1, 2, 3, 4, 5)]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 24

    by_slot = {(e.time_slot.day, e.time_slot.period): e for e in entries}
    max_period = max((p for _d, p in by_slot.keys()), default=8) if by_slot else 8

    for period in range(1, max_period + 1):
        row = 3 + period
        ws.cell(row=row, column=1, value=period).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 38
        for day in (1, 2, 3, 4, 5):
            cell = ws.cell(row=row, column=day + 1)
            entry = by_slot.get((day, period))
            if entry:
                if mode == 'class':
                    text = f'{entry.subject.name_he}\n{entry.teacher.first_name} {entry.teacher.last_name or ""}'.strip()
                else:
                    text = f'{entry.subject.name_he}\n{entry.school_class.display_name}'
                cell.value = text
                try:
                    cell.fill = PatternFill('solid', fgColor=entry.subject.color.lstrip('#'))
                except Exception:
                    pass
            cell.alignment = WRAP
            cell.font = GRID_FONT

    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 18 if c > 1 else 8


def _build_timetable_by_class(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.scheduling.models import TimetableEntry
    timetable_id = spec.get('timetable_id')
    if not timetable_id:
        return
    entries = list(
        TimetableEntry.objects.filter(timetable_id=timetable_id)
        .select_related('school_class', 'school_class__grade', 'subject', 'teacher', 'time_slot')
    )
    by_class: Dict[int, list] = defaultdict(list)
    class_names: Dict[int, str] = {}
    for e in entries:
        by_class[e.school_class_id].append(e)
        class_names[e.school_class_id] = e.school_class.display_name

    for cid in sorted(by_class.keys(), key=lambda i: class_names[i]):
        # Excel sheet names: max 31 chars, no special chars
        safe = class_names[cid][:25] + ' (כיתה)'
        ws = wb.create_sheet(safe[:31])
        _grid_for_entries(ws, f'מערכת שעות — כיתה {class_names[cid]}', by_class[cid], mode='class')


def _build_timetable_by_teacher(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.scheduling.models import TimetableEntry
    timetable_id = spec.get('timetable_id')
    if not timetable_id:
        return
    entries = list(
        TimetableEntry.objects.filter(timetable_id=timetable_id)
        .select_related('teacher', 'school_class', 'school_class__grade', 'subject', 'time_slot')
    )
    by_teacher: Dict[int, list] = defaultdict(list)
    teacher_names: Dict[int, str] = {}
    for e in entries:
        by_teacher[e.teacher_id].append(e)
        teacher_names[e.teacher_id] = f'{e.teacher.first_name} {e.teacher.last_name or ""}'.strip()

    for tid in sorted(by_teacher.keys(), key=lambda i: teacher_names[i]):
        safe = teacher_names[tid][:24] + ' (מורה)'
        ws = wb.create_sheet(safe[:31])
        _grid_for_entries(ws, f'מערכת שעות — {teacher_names[tid]}', by_teacher[tid], mode='teacher')


def _build_timetable_flat(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.scheduling.models import TimetableEntry
    timetable_id = spec.get('timetable_id')
    if not timetable_id:
        return
    entries = (
        TimetableEntry.objects.filter(timetable_id=timetable_id)
        .select_related('teacher', 'school_class', 'school_class__grade', 'subject', 'time_slot', 'room')
        .order_by('time_slot__day', 'time_slot__period', 'school_class__grade__level', 'school_class__number')
    )
    rows = [[
        e.id,
        e.time_slot.get_day_display(),
        e.time_slot.period,
        e.school_class.display_name,
        e.subject.name_he,
        f'{e.teacher.first_name} {e.teacher.last_name or ""}'.strip(),
        e.room.name if e.room else '',
    ] for e in entries]
    ws = wb.create_sheet('כל השיעורים')
    _write_table(ws, ['ID', 'יום', 'שיעור', 'כיתה', 'מקצוע', 'מורה', 'חדר'], rows)


def _build_conflicts(wb: Workbook, spec: Dict[str, Any]) -> None:
    """Detected conflicts in the selected timetable — same logic as the
    AI assistant's find_conflicts tool, but emitted as a sheet."""
    from apps.scheduling.models import TimetableEntry
    timetable_id = spec.get('timetable_id')
    if not timetable_id:
        return
    entries = list(
        TimetableEntry.objects.filter(timetable_id=timetable_id)
        .select_related('teacher', 'school_class', 'school_class__grade', 'subject', 'time_slot', 'room')
    )
    teacher_buckets, class_buckets, room_buckets = defaultdict(list), defaultdict(list), defaultdict(list)
    for e in entries:
        teacher_buckets[(e.time_slot_id, e.teacher_id)].append(e)
        class_buckets[(e.time_slot_id, e.school_class_id)].append(e)
        if e.room_id:
            room_buckets[(e.time_slot_id, e.room_id)].append(e)

    rows = []
    for es in teacher_buckets.values():
        if len(es) > 1:
            rows.append([
                'מורה כפול', f'{es[0].teacher.first_name} {es[0].teacher.last_name or ""}'.strip(),
                str(es[0].time_slot),
                ', '.join(f'{e.school_class.display_name}/{e.subject.name_he}' for e in es),
            ])
    for es in class_buckets.values():
        if len(es) > 1:
            rows.append([
                'כיתה כפולה', es[0].school_class.display_name,
                str(es[0].time_slot),
                ', '.join(f'{e.subject.name_he}/{e.teacher.first_name}' for e in es),
            ])
    for es in room_buckets.values():
        if len(es) > 1:
            rows.append([
                'חדר כפול', str(es[0].room),
                str(es[0].time_slot),
                ', '.join(f'{e.school_class.display_name}/{e.subject.name_he}' for e in es),
            ])
    ws = wb.create_sheet('התנגשויות')
    _write_table(ws, ['סוג התנגשות', 'ישות', 'משבצת', 'פרטים'], rows)


# ── audit / users (super-admin only — view layer enforces) ────────────────

def _build_audit_logins(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.users.models import AuditLogin
    qs = AuditLogin.objects.all().select_related('user')[:1000]
    rows = [[
        a.id, a.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        (a.user.email if a.user else '') or a.user_label,
        a.method, 'הצליח' if a.success else 'נכשל',
        a.ip_address or '', a.user_agent or '',
    ] for a in qs]
    ws = wb.create_sheet('יומן התחברויות')
    _write_table(ws, ['ID', 'מתי', 'משתמש', 'שיטה', 'סטטוס', 'IP', 'User-Agent'], rows)


def _build_audit_activities(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.users.models import AuditActivity
    qs = AuditActivity.objects.all().select_related('user')[:1000]
    rows = [[
        a.id, a.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        (a.user.email if a.user else '') or a.user_label,
        a.action, str(a.details or {}), a.ip_address or '',
    ] for a in qs]
    ws = wb.create_sheet('יומן פעולות')
    _write_table(ws, ['ID', 'מתי', 'משתמש', 'פעולה', 'פרטים', 'IP'], rows)


def _build_users(wb: Workbook, spec: Dict[str, Any]) -> None:
    from django.contrib.auth.models import User
    qs = User.objects.all().select_related('profile').order_by('id')
    rows = [[
        u.id, u.username, u.email,
        getattr(getattr(u, 'profile', None), 'full_name', '') or u.get_full_name(),
        getattr(getattr(u, 'profile', None), 'phone', '') or '',
        getattr(getattr(u, 'profile', None), 'role', ''),
        'כן' if u.is_active else 'לא',
        u.last_login.strftime('%Y-%m-%d %H:%M:%S') if u.last_login else '',
    ] for u in qs]
    ws = wb.create_sheet('משתמשים')
    _write_table(ws, ['ID', 'Username', 'אימייל', 'שם מלא', 'טלפון', 'תפקיד', 'פעיל', 'התחברות אחרונה'], rows)


def _build_import_logs(wb: Workbook, spec: Dict[str, Any]) -> None:
    from apps.import_export.models import ImportLog
    qs = ImportLog.objects.all().order_by('-uploaded_at')[:200]
    rows = [[
        log.id, log.uploaded_at.strftime('%Y-%m-%d %H:%M:%S'),
        log.file_name or '', log.get_status_display(),
        log.subjects_imported, log.teachers_imported, log.assignments_imported,
        '\n'.join(log.errors) if log.errors else '',
    ] for log in qs]
    ws = wb.create_sheet('יומן ייבואים')
    _write_table(ws, [
        'ID', 'מתי', 'שם קובץ', 'סטטוס',
        'מקצועות', 'מורים', 'שיבוצים', 'שגיאות',
    ], rows)


def _build_roundtrip_haarachot(wb: Workbook, spec: Dict[str, Any]) -> None:
    """Round-trip export: one sheet per subject in the school's original
    הערכות format. Lets the school cycle the data back to their Excel-
    based workflow after we've improved it in the UI.

    Each sheet matches the import format:
      - Row 1 col B: subject name
      - Row 3: header row (כיתה / סוג כיתה / שם המורה המלמד / שעות הוראה /
        שעות גמול בגרות / סמל שאלון בגרות / הערות)
      - Rows 4+: one row per active TeachingAssignment for the subject
    """
    from apps.subjects.models import Subject, TeachingAssignment
    school_id = spec.get('school_id') or 1

    # תפקידים sheet first.
    from apps.subjects.models import TeacherRole
    ws_roles = wb.create_sheet('תפקידים')
    ws_roles['A1'] = 'שם תפקיד'
    ws_roles['B1'] = 'הקשר'
    ws_roles['C1'] = 'תיאור'
    ws_roles['D1'] = 'מורה'
    ws_roles['E1'] = 'שעות שבועיות'
    ws_roles['F1'] = 'גמול תפקיד'
    ws_roles['G1'] = 'חובת הוראה'
    _style_header(ws_roles, 1, 7)
    role_row = 2
    for r in TeacherRole.objects.filter(school_id=school_id).select_related('teacher'):
        ws_roles.cell(row=role_row, column=1, value=r.role_title)
        ws_roles.cell(row=role_row, column=2, value=r.context)
        ws_roles.cell(row=role_row, column=3, value=r.description)
        ws_roles.cell(row=role_row, column=4, value=str(r.teacher) if r.teacher else '')
        ws_roles.cell(row=role_row, column=5, value=float(r.weekly_hours))
        ws_roles.cell(row=role_row, column=6, value=float(r.stipend_fraction))
        ws_roles.cell(row=role_row, column=7, value=float(r.must_teach_hours))
        role_row += 1
    _autosize(ws_roles)

    # One sheet per subject.
    for subject in Subject.objects.filter(school_id=school_id).order_by('name_he'):
        # Sheet titles must be unique and ≤ 31 chars; truncate if needed.
        sheet_name = subject.name_he[:31]
        if sheet_name in wb.sheetnames:
            continue
        ws = wb.create_sheet(sheet_name)
        ws['B1'] = subject.name_he
        ws['A3'] = 'כיתה'
        ws['B3'] = 'סוג כיתה'
        ws['C3'] = 'שם המורה המלמד'
        ws['D3'] = 'שעות הוראה'
        ws['E3'] = 'שעות גמול בגרות'
        ws['F3'] = 'סמל שאלון בגרות'
        ws['G3'] = 'הערות'
        _style_header(ws, 3, 7)
        row = 4
        for a in (
            TeachingAssignment.objects
            .filter(subject=subject)
            .select_related('teacher', 'school_class__grade')
            .prefetch_related('additional_classes')
            .order_by('school_class__grade__level', 'school_class__number')
        ):
            extras = list(a.additional_classes.values_list('grade__name', 'number'))
            if extras:
                cls_label = a.school_class.display_name + ',' + ','.join(
                    str(n) for _, n in extras
                )
            else:
                cls_label = a.school_class.display_name
            ws.cell(row=row, column=1, value=cls_label)
            ws.cell(row=row, column=2, value=a.track_label or '')
            ws.cell(row=row, column=3, value=str(a.teacher) if a.teacher else '')
            ws.cell(row=row, column=4, value=float(a.weekly_hours))
            ws.cell(row=row, column=5, value=float(a.bagrut_bonus_hours))
            ws.cell(row=row, column=6, value=a.bagrut_exam_code)
            ws.cell(row=row, column=7, value=a.notes)
            row += 1
        _autosize(ws)


def build_cleaned_workbook(parsed) -> Workbook:
    """Turn a messy uploaded workbook (already parsed into a ParsedWorkbook)
    into a clean, re-importable one — WITHOUT touching the database.

    For each subject it emits a standard sheet (header in row 3, one row per
    assignment), normalizing teacher names (canonical first+last) and class
    labels, carrying roles into a תפקידים sheet, and flagging every row that is
    missing a teacher (those won't be scheduled). A 'בדיקה' summary sheet lists
    the gaps to fix. The user fills in the blanks and re-imports.
    """
    from collections import OrderedDict
    from apps.import_export.parser import _canonicalize_teacher_name, _split_teacher_name

    wb = Workbook()
    default = wb.active

    SUBJECT_HEADERS = ['כיתה', 'סוג כיתה', 'שם המורה המלמד', 'שעות הוראה',
                       'שעות גמול בגרות', 'סמל שאלון בגרות', 'הערות']
    MISSING_FILL = PatternFill('solid', fgColor='FDE2E2')  # light red

    def _clean_teacher(raw: str) -> str:
        canon = _canonicalize_teacher_name(raw) if raw else ''
        if not canon:
            return ''
        first, last = _split_teacher_name(canon)
        return f'{first} {last}'.strip()

    def _class_label(classes) -> str:
        return ','.join(f'{g}{n}' for (g, n) in classes)

    by_subject: "OrderedDict[str, list]" = OrderedDict()
    for r in parsed.assignment_rows:
        by_subject.setdefault(r.subject or 'ללא מקצוע', []).append(r)

    missing: list[tuple[str, str]] = []   # (subject, class label)
    used_names: set[str] = set()

    for subject, rows in by_subject.items():
        name = (subject or 'מקצוע')[:31] or 'מקצוע'
        base, k = name, 2
        while name in used_names:
            name = f'{base[:28]}_{k}'
            k += 1
        used_names.add(name)
        ws = wb.create_sheet(name)
        ws['B1'] = subject
        ws['B1'].font = Font(bold=True, size=13)
        for c, h in enumerate(SUBJECT_HEADERS, start=1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row_i = 4
        for r in rows:
            teacher = _clean_teacher(r.teacher)
            note = r.notes or ''
            if not teacher:
                note = (note + ' | ' if note else '') + '⚠ חסר מורה — יש להשלים'
                missing.append((subject, _class_label(r.classes)))
            ws.cell(row=row_i, column=1, value=_class_label(r.classes))
            ws.cell(row=row_i, column=2, value=r.class_type_raw or '')
            tcell = ws.cell(row=row_i, column=3, value=teacher)
            if not teacher:
                tcell.fill = MISSING_FILL
            ws.cell(row=row_i, column=4, value=float(r.weekly_hours) if r.weekly_hours is not None else None)
            ws.cell(row=row_i, column=5, value=float(r.bagrut_hours) if r.bagrut_hours is not None else None)
            ws.cell(row=row_i, column=6, value=r.bagrut_code or '')
            ws.cell(row=row_i, column=7, value=note)
            row_i += 1
        _autosize(ws)

    # תפקידים
    if parsed.role_rows:
        ws = wb.create_sheet('תפקידים')
        role_headers = ['שם תפקיד', 'הקשר', 'תיאור', 'מורה', 'שעות שבועיות', 'גמול תפקיד', 'חובת הוראה']
        for c, h in enumerate(role_headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for i, rr in enumerate(parsed.role_rows, start=2):
            ws.cell(row=i, column=1, value=rr.role_title)
            ws.cell(row=i, column=2, value=rr.context)
            ws.cell(row=i, column=3, value=rr.description)
            ws.cell(row=i, column=4, value=_clean_teacher(rr.teacher))
            ws.cell(row=i, column=5, value=float(rr.weekly_hours) if rr.weekly_hours is not None else None)
            ws.cell(row=i, column=6, value=float(rr.stipend_fraction) if rr.stipend_fraction is not None else None)
            ws.cell(row=i, column=7, value=float(rr.must_teach_hours) if rr.must_teach_hours is not None else None)
        _autosize(ws)

    # בדיקה (summary of gaps) — placed first so the user sees it on open.
    summary = wb.create_sheet('בדיקה', 0)
    summary['A1'] = 'סיכום בדיקה — מה צריך להשלים'
    summary['A1'].font = Font(bold=True, size=14)
    summary['A3'] = 'מקצועות'
    summary['B3'] = len(by_subject)
    summary['A4'] = 'שורות שיבוץ'
    summary['B4'] = sum(len(v) for v in by_subject.values())
    summary['A5'] = 'שיעורים ללא מורה (להשלמה)'
    summary['B5'] = len(missing)
    summary['A5'].font = summary['B5'].font = Font(bold=True, color='B91C1C')
    summary['A7'] = 'מקצוע'
    summary['B7'] = 'כיתה'
    summary['A7'].font = summary['B7'].font = HEADER_FONT
    summary['A7'].fill = summary['B7'].fill = HEADER_FILL
    for i, (subj, cls) in enumerate(missing[:1000], start=8):
        summary.cell(row=i, column=1, value=subj)
        summary.cell(row=i, column=2, value=cls)
    _autosize(summary)

    if default.title == 'Sheet':
        wb.remove(default)
    return wb


# ── registry ──────────────────────────────────────────────────────────────
SHEET_BUILDERS: Dict[str, Callable[[Workbook, Dict[str, Any]], None]] = {
    'timetable_by_class': _build_timetable_by_class,
    'timetable_by_teacher': _build_timetable_by_teacher,
    'timetable_flat': _build_timetable_flat,
    'conflicts': _build_conflicts,
    'teachers': _build_teachers,
    'subjects': _build_subjects,
    'classes': _build_classes,
    'time_slots': _build_time_slots,
    'rooms': _build_rooms,
    'assignments': _build_assignments,
    'teacher_summary': _build_teacher_summary,
    'diagnostics': _build_diagnostics,
    'constraints': _build_constraints,
    'import_logs': _build_import_logs,
    'audit_logins': _build_audit_logins,        # super_admin only
    'audit_activities': _build_audit_activities,  # super_admin only
    'users': _build_users,                       # super_admin only
    'roundtrip_haarachot': _build_roundtrip_haarachot,
}

SUPER_ADMIN_ONLY_SHEETS = {'audit_logins', 'audit_activities', 'users'}


def build_workbook(spec: Dict[str, Any], is_super_admin: bool = False) -> Workbook:
    wb = Workbook()
    # openpyxl creates a default sheet; we'll remove it once we've added at
    # least one of our own.
    default = wb.active

    requested = spec.get('sheets') or []
    skipped: List[str] = []
    for name in requested:
        if name not in SHEET_BUILDERS:
            skipped.append(f'{name}: unknown sheet')
            continue
        if name in SUPER_ADMIN_ONLY_SHEETS and not is_super_admin:
            skipped.append(f'{name}: requires super_admin')
            continue
        SHEET_BUILDERS[name](wb, spec)

    # Drop the empty default sheet only if we successfully added others.
    if len(wb.worksheets) > 1 and default.title == 'Sheet':
        wb.remove(default)
    else:
        # Nothing was built — repurpose the default sheet for a friendly note
        default.title = 'מידע'
        default['A1'] = 'לא נבחרו לוחות לייצוא'
        default['A1'].font = Font(bold=True, size=12)
        for i, msg in enumerate(skipped, start=3):
            default.cell(row=i, column=1, value=msg)

    return wb
