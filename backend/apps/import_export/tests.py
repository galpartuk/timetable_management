"""
Tests for the Excel importer. Cover the parsing helpers and the end-to-end
analyze/apply flow against synthetic workbooks.
"""
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from openpyxl import Workbook

from apps.school.models import School, SchoolClass
from apps.subjects.models import Subject, Teacher, TeachingAssignment
from apps.import_export.parser import (
    _canonicalize_teacher_name,
    _parse_class_list,
    _parse_hours,
    _parse_single_class,
    analyze,
    apply,
)


class HelperTests(TestCase):

    def test_parse_single_class(self):
        self.assertEqual(_parse_single_class('ז1'), ('ז', 1))
        self.assertEqual(_parse_single_class('יא 3'), ('יא', 3))
        self.assertEqual(_parse_single_class("י' 7"), ('י', 7))
        self.assertEqual(_parse_single_class('יו"ד 5'), ('י', 5))
        self.assertEqual(_parse_single_class('י"א 10'), ('יא', 10))
        self.assertIsNone(_parse_single_class('הקבצה'))
        self.assertIsNone(_parse_single_class(''))
        self.assertIsNone(_parse_single_class(None))

    def test_parse_class_list(self):
        self.assertEqual(
            _parse_class_list('יא 1,2,5,6,7,9'),
            [('יא', 1), ('יא', 2), ('יא', 5), ('יא', 6), ('יא', 7), ('יא', 9)],
        )
        # The infamous "ח ,92,3,4,8" — typo for "9,2,3,4,8". Our heuristic
        # splits "92" into [9, 2] since class numbers don't exceed 12.
        self.assertEqual(
            sorted(_parse_class_list('ח ,92,3,4,8')),
            sorted([('ח', 9), ('ח', 2), ('ח', 3), ('ח', 4), ('ח', 8)]),
        )

    def test_parse_hours(self):
        self.assertEqual(_parse_hours(3), Decimal('3'))
        self.assertEqual(_parse_hours(3.5), Decimal('3.5'))
        self.assertEqual(_parse_hours('3+1'), Decimal('4'))
        self.assertEqual(_parse_hours('4+4'), Decimal('8'))
        self.assertIsNone(_parse_hours(''))
        self.assertIsNone(_parse_hours(None))

    def test_canonicalize_teacher_name(self):
        self.assertEqual(_canonicalize_teacher_name('אולגה עם ח7'), 'אולגה')
        self.assertEqual(_canonicalize_teacher_name('אולגה- ח,ט'), 'אולגה')
        self.assertEqual(_canonicalize_teacher_name('סוזנה- יא5'), 'סוזנה')
        self.assertEqual(_canonicalize_teacher_name('מאיה ק- יא3'), 'מאיה ק')
        self.assertEqual(_canonicalize_teacher_name('אגי-'), 'אגי')
        self.assertEqual(_canonicalize_teacher_name('?פאני'), 'פאני')
        self.assertEqual(_canonicalize_teacher_name('מיטל כהן (*2)'), 'מיטל כהן')


class EndToEndImportTests(TestCase):

    def _make_school(self):
        return School.objects.create(name='בית ספר לבדיקה')

    def _workbook_to_bytes(self, wb):
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _make_workbook(self, *, with_pool=False, with_roles=False, with_pe=False):
        wb = Workbook()
        ws = wb.active
        ws.title = 'מתמטיקה'
        ws['B1'] = 'מתמטיקה'
        ws['A3'] = 'כיתה'
        ws['B3'] = 'סוג כיתה'
        ws['C3'] = 'שם המורה המלמד'
        ws['D3'] = 'שעות הוראה'
        ws['G3'] = 'הערות'
        ws['A4'] = 'ז1'
        ws['B4'] = 'חינוך מיוחד'
        ws['C4'] = 'אולגה'
        ws['D4'] = 5
        ws['A5'] = 'ז2'
        ws['C5'] = 'דנה- ז2'
        ws['D5'] = 5

        if with_pool:
            ws['A6'] = 'יא 1,2,5'
            ws['B6'] = '5 יח"ל'
            ws['C6'] = 'מורן'
            ws['D6'] = 8
            ws['A7'] = ''
            ws['B7'] = '4 יח"ל'
            ws['C7'] = 'מיטל'
            ws['D7'] = 6

        if with_pe:
            ws_pe = wb.create_sheet('חינוך גופני')
            ws_pe['B1'] = 'חינוך גופני'
            ws_pe['A3'] = 'כיתה'
            ws_pe['B3'] = 'סוג כיתה'
            ws_pe['E3'] = 'שעות הוראה'
            ws_pe['F3'] = 'מורה לבנים'
            ws_pe['G3'] = 'מורה לבנות'
            ws_pe['A4'] = 'ז1'
            ws_pe['B4'] = 'חינוך מיוחד'
            ws_pe['E4'] = 2
            ws_pe['F4'] = 'שי'
            ws_pe['G4'] = 'אגי'

        if with_roles:
            ws_r = wb.create_sheet('תפקידים')
            ws_r['A1'] = 'שם תפקיד'
            ws_r['B1'] = 'הקשר'
            ws_r['C1'] = 'תיאור'
            ws_r['D1'] = 'מורה'
            ws_r['E1'] = 'שעות שבועיות'
            ws_r['F1'] = 'גמול תפקיד'
            ws_r['A2'] = 'מנהלת ביה"ס'
            ws_r['B2'] = 'חט"ע'
            ws_r['D2'] = 'ליאת'
            ws_r['E2'] = 24

        return wb

    def test_apply_creates_classes_and_subjects(self):
        school = self._make_school()
        wb = self._make_workbook()
        parsed = analyze(self._workbook_to_bytes(wb), 'test.xlsx')
        apply(parsed, school)

        self.assertEqual(Subject.objects.filter(school=school).count(), 1)
        self.assertEqual(SchoolClass.objects.filter(grade__school=school).count(), 2)
        z1 = SchoolClass.objects.get(grade__school=school, grade__name='ז', number=1)
        self.assertEqual(z1.class_type, SchoolClass.ClassType.SPECIAL_ED)
        teachers = list(Teacher.objects.filter(school=school).values_list('first_name', flat=True))
        self.assertIn('אולגה', teachers)
        self.assertIn('דנה', teachers)

    def test_pool_creates_one_assignment_with_extras(self):
        school = self._make_school()
        wb = self._make_workbook(with_pool=True)
        parsed = analyze(self._workbook_to_bytes(wb), 'test.xlsx')
        apply(parsed, school)

        pool_assignments = TeachingAssignment.objects.filter(
            subject__school=school,
            school_class__grade__name='יא',
            school_class__number=1,
            track_label__in=['5 יח"ל', '4 יח"ל'],
        )
        self.assertEqual(pool_assignments.count(), 2)
        for a in pool_assignments:
            members = a.additional_classes.values_list('grade__name', 'number')
            self.assertEqual(sorted(members), sorted([('יא', 2), ('יא', 5)]))

    def test_pe_sheet_produces_boys_and_girls_assignments(self):
        school = self._make_school()
        wb = self._make_workbook(with_pe=True)
        parsed = analyze(self._workbook_to_bytes(wb), 'test.xlsx')
        apply(parsed, school)

        pe_assignments = TeachingAssignment.objects.filter(
            subject__name_he='חינוך גופני', subject__school=school,
        )
        self.assertEqual(pe_assignments.count(), 2)
        tracks = sorted(pe_assignments.values_list('track_label', flat=True))
        self.assertEqual(tracks, sorted([
            'חינוך מיוחד / בנים', 'חינוך מיוחד / בנות',
        ]))

    def test_roles_sheet_imports_teacher_roles(self):
        from apps.subjects.models import TeacherRole
        school = self._make_school()
        wb = self._make_workbook(with_roles=True)
        parsed = analyze(self._workbook_to_bytes(wb), 'test.xlsx')
        apply(parsed, school)

        roles = TeacherRole.objects.filter(school=school)
        self.assertEqual(roles.count(), 1)
        r = roles.first()
        self.assertEqual(r.role_title, 'מנהלת ביה"ס')
        self.assertEqual(r.weekly_hours, Decimal('24'))

    def test_reimport_is_idempotent(self):
        school = self._make_school()
        wb = self._make_workbook()

        parsed1 = analyze(self._workbook_to_bytes(wb), 'test.xlsx')
        apply(parsed1, school)
        first_count = TeachingAssignment.objects.filter(subject__school=school).count()

        parsed2 = analyze(self._workbook_to_bytes(wb), 'test.xlsx')
        result2 = apply(parsed2, school)
        second_count = TeachingAssignment.objects.filter(subject__school=school).count()

        self.assertEqual(first_count, second_count)
        self.assertEqual(result2.assignments_created, 0)


class NormalizedExportTests(TestCase):
    """Phase 2 verification sheets: teacher summary + feasibility diagnostics."""

    def test_teacher_summary_and_diagnostics_flag_overload(self):
        from datetime import time
        from apps.school.models import Grade, School, SchoolClass, TimeSlot
        from apps.subjects.models import Subject, Teacher, TeachingAssignment
        from apps.import_export.exporter import build_workbook

        school = School.objects.create(name='בדיקת אבחון')
        grade = Grade.objects.create(school=school, name='ט', level=9)
        cls = SchoolClass.objects.create(grade=grade, number=4)
        # A four-slot week.
        for p in range(1, 5):
            TimeSlot.objects.create(
                school=school, day=1, period=p,
                start_time=time(8, 0), end_time=time(8, 45),
            )
        subj = Subject.objects.create(school=school, name_he='מתמטיקה')
        teacher = Teacher.objects.create(school=school, first_name='דנה')
        # 6 lessons but only 4 slots -> guaranteed class overload.
        TeachingAssignment.objects.create(
            subject=subj, teacher=teacher, school_class=cls, weekly_hours=Decimal('6'),
        )

        wb = build_workbook({'school_id': school.id, 'sheets': ['teacher_summary', 'diagnostics']})
        self.assertIn('סיכום מורים', wb.sheetnames)
        self.assertIn('אבחון', wb.sheetnames)

        summary_rows = list(wb['סיכום מורים'].iter_rows(values_only=True))
        self.assertGreaterEqual(len(summary_rows), 2)  # header + ≥1 data row
        self.assertIn('דנה', summary_rows[1])

        diag_codes = [r[1] for r in wb['אבחון'].iter_rows(values_only=True)]
        self.assertIn('class_overload', diag_codes)


class WipeEverythingTests(TestCase):
    """The full-reset bulk-delete op clears all imported/generated data but
    keeps the school and its structural setup (time slots)."""

    def test_wipe_everything(self):
        from datetime import time
        from django.contrib.auth.models import User
        from rest_framework.test import APIClient
        from apps.school.models import Grade, School, SchoolClass, TimeSlot
        from apps.scheduling.models import Timetable
        from apps.subjects.models import Subject, Teacher, TeachingAssignment

        school = School.objects.create(name='בדיקת איפוס')
        grade = Grade.objects.create(school=school, name='ז', level=7)
        cls = SchoolClass.objects.create(grade=grade, number=1)
        TimeSlot.objects.create(school=school, day=1, period=1,
                                start_time=time(8, 0), end_time=time(8, 45))
        subj = Subject.objects.create(school=school, name_he='מתמטיקה')
        teacher = Teacher.objects.create(school=school, first_name='דנה')
        TeachingAssignment.objects.create(subject=subj, teacher=teacher, school_class=cls,
                                          weekly_hours=Decimal('3'))
        Timetable.objects.create(school=school, name='tt')

        client = APIClient()
        client.force_authenticate(user=User.objects.create_user('admin2', password='x'))
        resp = client.post('/api/manage/bulk-delete/',
                           {'operation': 'wipe_everything', 'school_id': school.id},
                           format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(Teacher.objects.filter(school=school).count(), 0)
        self.assertEqual(Subject.objects.filter(school=school).count(), 0)
        self.assertEqual(TeachingAssignment.objects.filter(school_class__grade__school=school).count(), 0)
        self.assertEqual(SchoolClass.objects.filter(grade__school=school).count(), 0)
        self.assertEqual(Grade.objects.filter(school=school).count(), 0)
        self.assertEqual(Timetable.objects.filter(school=school).count(), 0)
        # Structural setup is preserved.
        self.assertEqual(TimeSlot.objects.filter(school=school).count(), 1)
        self.assertTrue(School.objects.filter(id=school.id).exists())


class TeacherNameSplitTests(TestCase):
    def test_split(self):
        from apps.import_export.parser import _split_teacher_name
        self.assertEqual(_split_teacher_name('אורית'), ('אורית', ''))
        self.assertEqual(_split_teacher_name('אורית יחזקאל'), ('אורית', 'יחזקאל'))
        self.assertEqual(_split_teacher_name('מאיה ק'), ('מאיה ק', ''))  # single-letter suffix kept
        self.assertEqual(_split_teacher_name('רינת שוורץ'), ('רינת', 'שוורץ'))


class TeacherResolveTests(TestCase):
    def setUp(self):
        self.school = School.objects.create(name='בדיקת זיהוי')

    def _idx(self):
        from apps.import_export.parser import _build_existing_index
        return _build_existing_index(self.school)

    def _resolve(self, name):
        from apps.import_export.parser import resolve_teacher_match
        return resolve_teacher_match(name, self._idx())

    def test_exact_match(self):
        t = Teacher.objects.create(school=self.school, first_name='אורית', last_name='יחזקאל')
        res = self._resolve('אורית יחזקאל')
        self.assertEqual(res['status'], 'matched')
        self.assertEqual(res['suggested'], t.id)

    def test_first_only_one_full_candidate_is_ambiguous_suggest_merge(self):
        t = Teacher.objects.create(school=self.school, first_name='אורית', last_name='יחזקאל')
        res = self._resolve('אורית')
        self.assertEqual(res['status'], 'ambiguous')
        self.assertEqual(res['suggested'], t.id)

    def test_two_candidates_ambiguous_suggest_new(self):
        Teacher.objects.create(school=self.school, first_name='אורית', last_name='יחזקאל')
        Teacher.objects.create(school=self.school, first_name='אורית', last_name='כהן')
        res = self._resolve('אורית')
        self.assertEqual(res['status'], 'ambiguous')
        self.assertEqual(res['suggested'], 'new')

    def test_no_match_is_new(self):
        self.assertEqual(self._resolve('יוסי')['status'], 'new')

    def test_legacy_full_name_in_first_name_indexes_and_matches(self):
        t = Teacher.objects.create(school=self.school, first_name='אורית יחזקאל', last_name='')
        res = self._resolve('אורית')
        self.assertEqual(res['suggested'], t.id)
        self.assertEqual(res['status'], 'ambiguous')


class TeacherForOverrideTests(TestCase):
    def setUp(self):
        self.school = School.objects.create(name='בדיקת עקיפה')

    def _call(self, name, **kw):
        from apps.import_export.parser import _teacher_for, _build_existing_index
        kw.setdefault('existing_index', _build_existing_index(self.school))
        return _teacher_for(self.school, name, {}, **kw)

    def test_override_to_existing_id(self):
        t = Teacher.objects.create(school=self.school, first_name='אורית', last_name='יחזקאל')
        got = self._call('אורית', overrides={'אורית': t.id})
        self.assertEqual(got.id, t.id)
        self.assertEqual(Teacher.objects.filter(school=self.school).count(), 1)

    def test_override_new_splits_name(self):
        Teacher.objects.create(school=self.school, first_name='אורית', last_name='יחזקאל')
        got = self._call('אורית כהן', overrides={'אורית כהן': 'new'})
        self.assertEqual((got.first_name, got.last_name), ('אורית', 'כהן'))
        self.assertEqual(Teacher.objects.filter(school=self.school).count(), 2)

    def test_auto_merge_backfills_empty_last_name(self):
        t = Teacher.objects.create(school=self.school, first_name='אורית', last_name='')
        got = self._call('אורית יחזקאל')
        self.assertEqual(got.id, t.id)
        got.refresh_from_db()
        self.assertEqual(got.last_name, 'יחזקאל')

    def test_bad_override_falls_back_to_auto(self):
        warns = []
        got = self._call('יוסי', overrides={'יוסי': 9999}, warnings=warns)
        self.assertIsNotNone(got)
        self.assertTrue(warns)


class ComputeTeacherResolutionsTests(TestCase):
    def test_ambiguous_surfaced(self):
        import types
        from apps.import_export.views import _compute_teacher_resolutions
        school = School.objects.create(name='בדיקת תצוגה')
        Teacher.objects.create(school=school, first_name='אורית', last_name='יחזקאל')
        parsed = types.SimpleNamespace(
            assignment_rows=[types.SimpleNamespace(teacher='אורית')],
            role_rows=[],
        )
        out = _compute_teacher_resolutions(parsed, school)
        self.assertEqual(out['ambiguous_count'], 1)
        entry = out['ambiguous'][0]
        self.assertEqual(entry['incoming'], 'אורית')
        # choices = the candidate + a "new" option
        values = [c['value'] for c in entry['choices']]
        self.assertIn('new', values)
