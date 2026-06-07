import io
from collections import Counter

from django.core.files.base import ContentFile
from django.db import transaction
from django.http import FileResponse, HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.school.models import School, SchoolClass
from apps.scheduling.models import Timetable, TimetableEntry
from apps.subjects.models import Subject, Teacher, TeachingAssignment

from .exporter import SHEET_BUILDERS, SUPER_ADMIN_ONLY_SHEETS, build_workbook
from .models import ImportLog
from .parser import analyze, apply as apply_parsed, parse_timetable_excel
from .parser_days_off import parse_days_off_excel


def _is_super_admin(user) -> bool:
    return getattr(getattr(user, 'profile', None), 'role', None) == 'super_admin'


def _diff_against_db(parsed, school) -> dict:
    """Compare a fresh ParsedWorkbook against the current DB state for
    one school. Returns counts and small samples of:

      • new_teachers / removed_teachers
      • new_classes
      • new_subjects
      • assignments with changed weekly_hours (key on sheet+row)
      • assignments in the import not in the DB (will be added)
      • assignments in the DB not in the import (will linger as stale)

    Used by the dry-run preview screen — shows the user exactly what
    is going to change before they commit.
    """
    from apps.subjects.models import Teacher, Subject, TeachingAssignment
    from apps.school.models import SchoolClass

    from .parser import _valid_canonical, _build_existing_index, resolve_teacher_match

    db_subjects = set(Subject.objects.filter(school=school)
                      .values_list('name_he', flat=True))
    db_class_keys = set(
        SchoolClass.objects.filter(grade__school=school)
        .values_list('grade__name', 'number')
    )

    import_subjects: set[str] = set()
    import_class_keys: set[tuple[str, int]] = set()
    import_canon: set[str] = set()
    for r in parsed.assignment_rows:
        c = _valid_canonical(r.teacher) if r.teacher else None
        if c:
            import_canon.add(c)
        import_subjects.add(r.subject)
        for g, n in r.classes:
            import_class_keys.add((g, n))
    for rr in parsed.role_rows:
        c = _valid_canonical(rr.teacher) if rr.teacher else None
        if c:
            import_canon.add(c)

    # New vs matched teachers via the same first+last resolver used on commit,
    # so the counts match what the import will actually do.
    existing_index = _build_existing_index(school)
    matched_ids: set[int] = set()
    new_teachers: list[str] = []
    for c in sorted(import_canon):
        res = resolve_teacher_match(c, existing_index)
        if res['status'] == 'new':
            new_teachers.append(c)
        elif res['suggested'] != 'new':
            matched_ids.add(res['suggested'])
    removed_teachers = sorted(
        str(t) for t in Teacher.objects.filter(school=school) if t.id not in matched_ids
    )

    # Existing assignments keyed by (sheet, row) so we can compare hours.
    db_hours_by_source = {
        (a.source_sheet, a.source_row): float(a.weekly_hours or 0)
        for a in TeachingAssignment.objects.filter(subject__school=school)
        if a.source_sheet and a.source_row
    }
    hours_changes = []
    new_rows_seen = set()
    for r in parsed.assignment_rows:
        if r.weekly_hours is None:
            continue
        key = (r.source_sheet, r.source_row)
        new_rows_seen.add(key)
        existing = db_hours_by_source.get(key)
        new_hours = float(r.weekly_hours)
        if existing is None:
            continue  # truly new row → covered by `new_rows`
        if abs(existing - new_hours) > 0.01:
            hours_changes.append({
                'sheet': r.source_sheet, 'row': r.source_row,
                'teacher': r.teacher, 'subject': r.subject,
                'old_hours': existing, 'new_hours': new_hours,
            })

    new_rows = [k for k in new_rows_seen if k not in db_hours_by_source]
    stale_rows = [k for k in db_hours_by_source if k not in new_rows_seen]

    return {
        'new_teachers': new_teachers[:50],
        'new_teachers_count': len(new_teachers),
        'removed_teachers': removed_teachers[:50],
        'removed_teachers_count': len(removed_teachers),
        'new_subjects': sorted(import_subjects - db_subjects)[:30],
        'new_classes': sorted([f'{g}{n}' for (g, n) in import_class_keys - db_class_keys])[:30],
        'new_rows_count': len(new_rows),
        'stale_rows_count': len(stale_rows),
        'hours_changes_count': len(hours_changes),
        'hours_changes': hours_changes[:30],
    }


def _compute_teacher_resolutions(parsed, school) -> dict:
    """Identify incoming teacher names whose identity is ambiguous (could be an
    existing teacher OR a new person) so the import review step can ask the
    user. Names that match cleanly or are clearly new are resolved silently.

    Returns ``{ambiguous: [{incoming, first, last, candidates, suggested,
    choices}], ambiguous_count, auto_matched_count}``. ``choices`` carries the
    radio options (each candidate id + a 'new' option); the override map sent
    back on commit is ``{incoming_name: id | 'new'}``.
    """
    from .parser import _valid_canonical, _build_existing_index, resolve_teacher_match

    existing_index = _build_existing_index(school)
    seen: set[str] = set()
    canon_names: list[str] = []
    for r in list(parsed.assignment_rows) + list(parsed.role_rows):
        raw = getattr(r, 'teacher', None)
        c = _valid_canonical(raw) if raw else None
        if c and c not in seen:
            seen.add(c)
            canon_names.append(c)

    ambiguous = []
    auto_matched = 0
    for c in canon_names:
        res = resolve_teacher_match(c, existing_index)
        if res['status'] == 'ambiguous':
            choices = [
                {'value': cand['id'], 'label': f"מיזוג עם {cand['display_name']}"}
                for cand in res['candidates']
            ]
            choices.append({'value': 'new', 'label': 'צור מורה חדש'})
            ambiguous.append({
                'incoming': c, 'first': res['first'], 'last': res['last'],
                'candidates': res['candidates'], 'suggested': res['suggested'],
                'choices': choices,
            })
        elif res['status'] == 'matched':
            auto_matched += 1

    return {
        'ambiguous': ambiguous,
        'ambiguous_count': len(ambiguous),
        'auto_matched_count': auto_matched,
    }


def _compute_coverage(parsed) -> dict:
    """Estimate how full the generated timetable will be, per class, from the
    parsed rows — BEFORE committing. A lesson is "schedulable" only if its row
    has a teacher; rows with a blank teacher cell can't be placed by the solver
    and show up as a gap. Gives the user a rough "this is how it'll look" view
    so missing teachers are caught at import time, not after generating.

    Only active rows with hours > 0 count. Pooled/track rows are counted per
    class, so totals are an estimate, not an exact slot count.
    """
    from collections import defaultdict
    from .parser import _valid_canonical

    scheduled: dict[str, float] = defaultdict(float)
    missing: dict[str, float] = defaultdict(float)
    for r in parsed.assignment_rows:
        if not r.is_active:
            continue
        hrs = float(r.weekly_hours or 0)
        if hrs <= 0:
            continue
        has_teacher = bool(_valid_canonical(r.teacher)) if r.teacher else False
        for (g, n) in r.classes:
            key = f'{g}{n}'
            (scheduled if has_teacher else missing)[key] += hrs

    classes = []
    for key in set(scheduled) | set(missing):
        s = round(scheduled.get(key, 0), 1)
        m = round(missing.get(key, 0), 1)
        classes.append({
            'class': key, 'scheduled_hours': s, 'missing_hours': m,
            'total_hours': round(s + m, 1),
        })
    # Worst gaps first so the user sees the problem classes at the top.
    classes.sort(key=lambda c: (-c['missing_hours'], c['class']))
    return {
        'classes': classes,
        'total_scheduled': round(sum(scheduled.values()), 1),
        'total_missing': round(sum(missing.values()), 1),
        'classes_with_gaps': sum(1 for c in classes if c['missing_hours'] > 0),
    }


def _compute_assignment_preview(parsed, limit: int = 3000) -> dict:
    """Flat 'who teaches what to whom' list from the parsed rows, so the user
    can verify the class → subject → teacher mapping before committing. Pooled
    rows are expanded to one entry per class. The teacher is the canonical name
    the import will set (blank = no teacher in the file)."""
    from .parser import _valid_canonical

    rows = []
    for r in parsed.assignment_rows:
        teacher = (_valid_canonical(r.teacher) or '') if r.teacher else ''
        for (g, n) in r.classes:
            rows.append({
                'class': f'{g}{n}',
                'subject': r.subject,
                'teacher': teacher,
                'hours': float(r.weekly_hours or 0),
                'active': bool(r.is_active),
            })
    rows.sort(key=lambda x: (x['class'], x['subject'], x['teacher']))
    return {'rows': rows[:limit], 'total': len(rows), 'truncated': len(rows) > limit}


def _summarize_parsed(parsed) -> dict:
    """Build a digest shown in the dry-run preview screen."""
    subjects = Counter(r.subject for r in parsed.assignment_rows)
    teachers = Counter()
    class_grade = Counter()
    rows_with_teacher = 0
    rows_with_hours = 0
    pool_rows = 0
    inactive_rows = 0
    for r in parsed.assignment_rows:
        if r.teacher:
            teachers[r.teacher] += 1
            rows_with_teacher += 1
        if r.weekly_hours:
            rows_with_hours += 1
        if len(r.classes) > 1:
            pool_rows += 1
        if not r.is_active:
            inactive_rows += 1
        for g, n in r.classes:
            class_grade[g] += 1

    role_teachers = Counter(r.teacher for r in parsed.role_rows if r.teacher)

    return {
        'sheets_seen': parsed.sheets_seen,
        'assignment_rows_total': len(parsed.assignment_rows),
        'role_rows_total': len(parsed.role_rows),
        'subjects_distinct': len(subjects),
        'teachers_distinct': len(teachers | role_teachers),
        'rows_with_teacher': rows_with_teacher,
        'rows_with_hours': rows_with_hours,
        'pool_rows': pool_rows,
        'inactive_rows': inactive_rows,
        'class_rows_per_grade': dict(class_grade),
        'top_subjects': subjects.most_common(15),
        'top_teachers': teachers.most_common(15),
        'warnings': parsed.warnings,
        'errors': parsed.errors,
    }


@api_view(['POST'])
@parser_classes([MultiPartParser])
def upload_excel(request):
    """Parse + commit. If ``dry_run=true``, only the preview is built and
    the row is stored with status=PREVIEW for the FE to confirm later.

    Body (multipart):
      file: the .xlsx
      school_id: int
      dry_run: 'true' / 'false' (default false)
      wipe_existing: 'true' / 'false' (default false) — clears all teaching
        data for the school before importing. Use sparingly.
    """
    file = request.FILES.get('file')
    school_id = request.data.get('school_id')
    dry_run = (request.data.get('dry_run') or '').lower() == 'true'
    wipe = (request.data.get('wipe_existing') or '').lower() == 'true'
    # User identity decisions from the review step: {canonical_name: id | 'new'}.
    overrides_raw = request.data.get('teacher_overrides')
    teacher_overrides = None
    if overrides_raw:
        import json
        try:
            teacher_overrides = json.loads(overrides_raw)
        except (ValueError, TypeError):
            teacher_overrides = None

    if not file:
        return Response({'error': 'לא נבחר קובץ'}, status=status.HTTP_400_BAD_REQUEST)
    if not school_id:
        return Response({'error': 'לא נבחר בית ספר'}, status=status.HTTP_400_BAD_REQUEST)
    school = School.objects.filter(id=school_id).first()
    if not school:
        return Response({'error': f'בית ספר {school_id} לא נמצא'}, status=status.HTTP_404_NOT_FOUND)

    import_log = ImportLog.objects.create(
        school=school,
        file_name=file.name,
        status=ImportLog.Status.PROCESSING,
        is_dry_run=dry_run,
    )

    # Keep the raw bytes so we can persist the source file on commit (analyze
    # consumes the stream, so read it up front and rewind for the parser).
    raw_bytes = file.read()
    file.seek(0)

    try:
        # Phase 1: analyze (no DB writes).
        parsed = analyze(file, file_name=file.name)
        preview = _summarize_parsed(parsed)
        preview['diff'] = _diff_against_db(parsed, school)
        preview['teacher_resolutions'] = _compute_teacher_resolutions(parsed, school)
        preview['coverage'] = _compute_coverage(parsed)
        preview['assignments_preview'] = _compute_assignment_preview(parsed)
        import_log.preview_data = preview
        import_log.warnings = parsed.warnings
        import_log.errors = parsed.errors

        if dry_run:
            import_log.status = ImportLog.Status.PREVIEW
            import_log.save()
            return Response({
                'message': 'תצוגה מקדימה הוכנה — אישור נדרש לייבוא בפועל',
                'log_id': import_log.id,
                'preview': preview,
                'dry_run': True,
            })

        # Phase 2: commit.
        result = apply_parsed(parsed, school, wipe_existing=wipe, teacher_overrides=teacher_overrides)
        import_log.subjects_imported = result.subjects_created
        import_log.teachers_imported = result.teachers_created
        import_log.classes_imported = result.classes_created
        import_log.assignments_imported = result.assignments_created + result.assignments_updated
        import_log.roles_imported = result.roles_created
        import_log.warnings = result.warnings
        import_log.errors = result.errors
        import_log.details = {
            'assignments_created': result.assignments_created,
            'assignments_updated': result.assignments_updated,
            'sheets_seen': parsed.sheets_seen,
        }
        # Persist the source file so the user can see/re-download the Excel
        # behind the currently-loaded data.
        import_log.source_file.save(file.name, ContentFile(raw_bytes), save=False)
        import_log.status = ImportLog.Status.COMPLETED
        import_log.save()
        return Response({
            'message': 'הייבוא הושלם בהצלחה',
            'log_id': import_log.id,
            'subjects_imported': import_log.subjects_imported,
            'teachers_imported': import_log.teachers_imported,
            'classes_imported': import_log.classes_imported,
            'assignments_imported': import_log.assignments_imported,
            'roles_imported': import_log.roles_imported,
            'warnings': import_log.warnings,
            'errors': import_log.errors,
            'preview': preview,
        })
    except Exception as e:
        import_log.status = ImportLog.Status.FAILED
        import_log.errors = [str(e)]
        import_log.save()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@parser_classes([MultiPartParser])
def upload_days_off(request):
    file = request.FILES.get('file')
    school_id = request.data.get('school_id')

    if not file:
        return Response({'error': 'לא נבחר קובץ'}, status=status.HTTP_400_BAD_REQUEST)
    if not school_id:
        return Response({'error': 'לא נבחר בית ספר'}, status=status.HTTP_400_BAD_REQUEST)

    school = School.objects.get(id=school_id)
    import_log = ImportLog.objects.create(
        school=school,
        file_name=file.name,
        status=ImportLog.Status.PROCESSING,
    )

    try:
        result = parse_days_off_excel(file, school, import_log)
        import_log.status = ImportLog.Status.COMPLETED
        import_log.save()
        return Response({
            'message': 'הייבוא הושלם בהצלחה',
            'teachers_updated': result['teachers_updated'],
            'errors': result['errors'],
        })
    except Exception as e:
        import_log.status = ImportLog.Status.FAILED
        import_log.errors = [str(e)]
        import_log.save()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def import_logs(request):
    school_id = request.query_params.get('school_id')
    logs = ImportLog.objects.filter(school_id=school_id) if school_id else ImportLog.objects.all()
    data = [{
        'id': log.id,
        'file_name': log.file_name,
        'status': log.status,
        'is_dry_run': log.is_dry_run,
        'uploaded_at': log.uploaded_at,
        'subjects_imported': log.subjects_imported,
        'teachers_imported': log.teachers_imported,
        'classes_imported': log.classes_imported,
        'assignments_imported': log.assignments_imported,
        'roles_imported': log.roles_imported,
        'warnings': log.warnings,
        'errors': log.errors,
        'has_file': bool(log.source_file),
    } for log in logs[:20]]
    return Response(data)


@api_view(['GET'])
def current_data_source(request):
    """Summary of the data currently loaded for a school: the most recent
    successful import (file name + when), whether its source .xlsx is stored,
    and LIVE counts from the DB (not the import-time counts — the data may have
    changed since). Drives the "loaded data" panel on the timetable page."""
    school_id = request.query_params.get('school_id')
    if not school_id:
        return Response({'error': 'school_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    last = (
        ImportLog.objects
        .filter(school_id=school_id, status=ImportLog.Status.COMPLETED, is_dry_run=False)
        .order_by('-uploaded_at')
        .first()
    )

    counts = {
        'teachers': Teacher.objects.filter(school_id=school_id).count(),
        'classes': SchoolClass.objects.filter(grade__school_id=school_id).count(),
        'subjects': Subject.objects.filter(school_id=school_id).count(),
        'assignments': TeachingAssignment.objects.filter(subject__school_id=school_id).count(),
    }
    counts['has_data'] = any(counts.values())

    source = None
    if last:
        source = {
            'log_id': last.id,
            'file_name': last.file_name,
            'uploaded_at': last.uploaded_at,
            'has_file': bool(last.source_file),
        }
    return Response({'source': source, 'counts': counts})


@api_view(['GET'])
def download_import_file(request, log_id):
    """Stream back the stored source .xlsx for an import log."""
    log = ImportLog.objects.filter(id=log_id).first()
    if not log or not log.source_file:
        return Response({'error': 'הקובץ לא נמצא'}, status=status.HTTP_404_NOT_FOUND)
    return FileResponse(
        log.source_file.open('rb'), as_attachment=True, filename=log.file_name,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def import_template(request):
    """Generate and stream back a blank Excel template with the columns the
    parser expects, plus two example rows. Built at request time (rather than
    shipped as a static file) so changes to HEADER_KEYS don't silently drift
    out of sync with the template a user downloads."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'מתמטיקה'  # one example subject sheet
    headers = [
        'כיתה', 'סוג כיתה', 'שם המורה המלמד', 'שעות הוראה',
        'שעות גמול בגרות', 'סמל שאלון בגרות', 'הערות',
    ]
    ws.append(headers)
    # Header styling — visually distinct so the user knows not to delete the row.
    bold = Font(bold=True)
    fill = PatternFill(fill_type='solid', fgColor='E0E7FF')
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='center')
    # Example rows so the user sees the expected shape (RTL Hebrew).
    ws.append(['ז1', '', 'דנה כהן', 5, '', '', 'דוגמה — מחקו את השורות לדוגמה לפני ייבוא'])
    ws.append(['ז2', '', 'דנה כהן', 5, '', '', ''])
    ws.append(['יא 1,2,5', '', '', '', '', '', 'שורת הקבצה — בשורות הבאות פירוט יח"ל/מורה'])
    ws.append(['', '5 יח"ל', 'משה לוי', 8, 2, '035001', ''])
    ws.append(['', '4 יח"ל', 'רינה דוד', 6, 1, '035001', ''])
    # Reasonable column widths so Hebrew text is readable.
    widths = [16, 14, 22, 12, 14, 14, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.sheet_view.rightToLeft = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="timetable_template.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def gap_analysis(request):
    """Snapshot of "what's missing for a feasible timetable".

    Returns counts and lists of:
      • classes with no homeroom teacher
      • subject deliveries with no teacher (TBD)
      • subject deliveries with no hours
      • teacher hour load vs their cap
      • teacher hours that overlap their role's must_teach floor

    The FE shows this on the Manage page so the school can fix the gaps
    before running the solver.
    """
    from apps.school.models import SchoolClass
    from apps.subjects.models import Teacher, TeacherRole
    from django.db.models import Sum, Count

    school_id = int(request.query_params.get('school_id') or 1)

    classes_missing_homeroom = list(
        SchoolClass.objects.filter(
            grade__school_id=school_id, homeroom_teacher__isnull=True,
        ).select_related('grade').values('id', 'grade__name', 'number')
    )

    assignments_without_teacher = list(
        TeachingAssignment.objects.filter(
            subject__school_id=school_id, teacher__isnull=True, is_active=True,
        ).select_related('subject', 'school_class__grade').values(
            'id', 'subject__name_he', 'school_class__grade__name',
            'school_class__number', 'weekly_hours', 'track_label',
        )[:200]
    )

    assignments_without_hours = list(
        TeachingAssignment.objects.filter(
            subject__school_id=school_id, weekly_hours=0,
        ).values('id', 'subject__name_he', 'school_class__grade__name',
                 'school_class__number')[:200]
    )

    from django.db.models import Q
    teacher_loads = []
    # Only active assignments count toward the load — inactive ones are
    # "פתיחה מותנית" (conditional) and shouldn't pressure the cap.
    for t in Teacher.objects.filter(school_id=school_id).annotate(
        load=Sum('assignments__weekly_hours', filter=Q(assignments__is_active=True)),
        role_hours=Sum('roles__weekly_hours'),
        must_teach=Sum('roles__must_teach_hours'),
    ).order_by('-load'):
        teacher_loads.append({
            'id': t.id,
            'name': str(t),
            'assigned_hours': float(t.load or 0),
            'role_hours': float(t.role_hours or 0),
            'must_teach': float(t.must_teach or 0),
            'cap': t.max_weekly_hours,
            'over_cap': (float(t.load or 0) > t.max_weekly_hours),
            'under_must_teach': (
                float(t.must_teach or 0) > 0
                and float(t.load or 0) < float(t.must_teach or 0)
            ),
        })

    return Response({
        'school_id': school_id,
        'classes_missing_homeroom': classes_missing_homeroom,
        'classes_missing_homeroom_count': len(classes_missing_homeroom),
        'assignments_without_teacher': assignments_without_teacher,
        'assignments_without_teacher_count': len(assignments_without_teacher),
        'assignments_without_hours': assignments_without_hours,
        'assignments_without_hours_count': len(assignments_without_hours),
        'teacher_loads': teacher_loads,
    })


# ─────────────────────────────────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_export_options(request):
    """Tells the FE which sheets are available + which require super-admin.
    Used to render the checkbox grid."""
    is_sa = _is_super_admin(request.user)
    return Response({
        'sheets': sorted(SHEET_BUILDERS.keys()),
        'super_admin_only': sorted(SUPER_ADMIN_ONLY_SHEETS),
        'is_super_admin': is_sa,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_excel(request):
    """Build a multi-sheet xlsx based on the spec and stream it back.

    Body: {school_id?, timetable_id?, sheets: [..]}
    Response: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    """
    spec = request.data or {}
    sheets = spec.get('sheets') or []
    if not isinstance(sheets, list) or not sheets:
        return Response({'error': 'sheets must be a non-empty list'},
                        status=status.HTTP_400_BAD_REQUEST)

    is_sa = _is_super_admin(request.user)
    # Strip super-admin-only sheets early so non-admins get a friendlier error
    # before we even start building the workbook.
    forbidden = [s for s in sheets if s in SUPER_ADMIN_ONLY_SHEETS and not is_sa]
    if forbidden:
        return Response(
            {'error': f'super_admin required for: {", ".join(forbidden)}'},
            status=status.HTTP_403_FORBIDDEN,
        )

    wb = build_workbook(spec, is_super_admin=is_sa)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timetable_id = spec.get('timetable_id')
    filename = f'timetable_export_{timetable_id or "all"}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────
# BULK DELETE
# ─────────────────────────────────────────────────────────────────────────
# Each operation is its own endpoint so they can be permission-checked
# individually and the FE shows a clear "this will delete X" preview.

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_timetable(request, timetable_id: int):
    """Delete a single timetable. Cascades to its entries via the FK."""
    tt = Timetable.objects.filter(id=timetable_id).first()
    if not tt:
        return Response({'error': 'not found'}, status=status.HTTP_404_NOT_FOUND)
    name = tt.name
    entry_count = TimetableEntry.objects.filter(timetable=tt).count()
    tt.delete()
    return Response({'ok': True, 'deleted': name, 'entries_deleted': entry_count})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def clear_timetable_entries(request, timetable_id: int):
    """Wipe every TimetableEntry on a timetable but keep the timetable
    record (so the user can re-run the generator on it)."""
    tt = Timetable.objects.filter(id=timetable_id).first()
    if not tt:
        return Response({'error': 'not found'}, status=status.HTTP_404_NOT_FOUND)
    deleted, _ = TimetableEntry.objects.filter(timetable=tt).delete()
    tt.status = Timetable.Status.DRAFT
    tt.save(update_fields=['status'])
    return Response({'ok': True, 'entries_deleted': deleted})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_delete(request):
    """Multi-purpose destructive operations. Body: {operation, school_id?}.

    Operations:
      - clear_assignments      remove all TeachingAssignments for school
      - clear_all_timetables   remove every Timetable for school (cascades)
      - clear_subjects         remove all Subjects (only if not referenced)
      - clear_teachers         remove all Teachers (only if not referenced)
      - wipe_school_data       super-admin only — nukes timetables,
                               assignments, constraints, entries. Keeps
                               teachers, subjects, classes, time slots.
    """
    op = (request.data or {}).get('operation')
    school_id = (request.data or {}).get('school_id') or 1
    if not School.objects.filter(id=school_id).exists():
        return Response({'error': f'school {school_id} not found'},
                        status=status.HTTP_404_NOT_FOUND)

    is_sa = _is_super_admin(request.user)
    summary: dict = {}

    if op == 'clear_assignments':
        n, _ = TeachingAssignment.objects.filter(school_class__grade__school_id=school_id).delete()
        summary['assignments_deleted'] = n

    elif op == 'clear_all_timetables':
        # Cascade handles entries.
        with transaction.atomic():
            entries = TimetableEntry.objects.filter(timetable__school_id=school_id).count()
            tts, _ = Timetable.objects.filter(school_id=school_id).delete()
            summary['timetables_deleted'] = tts
            summary['entries_deleted'] = entries

    elif op == 'clear_subjects':
        # Refuse if any subject is referenced by an assignment or entry —
        # that would silently nuke history. Make the user clear those first.
        if (
            TeachingAssignment.objects.filter(subject__school_id=school_id).exists()
            or TimetableEntry.objects.filter(subject__school_id=school_id).exists()
        ):
            return Response({
                'error': 'subjects are referenced by assignments or timetable entries; '
                         'clear those first',
            }, status=status.HTTP_409_CONFLICT)
        n, _ = Subject.objects.filter(school_id=school_id).delete()
        summary['subjects_deleted'] = n

    elif op == 'clear_teachers':
        if (
            TeachingAssignment.objects.filter(teacher__school_id=school_id).exists()
            or TimetableEntry.objects.filter(teacher__school_id=school_id).exists()
        ):
            return Response({
                'error': 'teachers are referenced by assignments or timetable entries; '
                         'clear those first',
            }, status=status.HTTP_409_CONFLICT)
        n, _ = Teacher.objects.filter(school_id=school_id).delete()
        summary['teachers_deleted'] = n

    elif op == 'wipe_school_data':
        if not is_sa:
            return Response({'error': 'super_admin required'},
                            status=status.HTTP_403_FORBIDDEN)
        with transaction.atomic():
            from apps.scheduling.models import Constraint
            entries = TimetableEntry.objects.filter(timetable__school_id=school_id).count()
            tts = Timetable.objects.filter(school_id=school_id).count()
            asn = TeachingAssignment.objects.filter(school_class__grade__school_id=school_id).count()
            cons = Constraint.objects.filter(school_id=school_id).count()
            Timetable.objects.filter(school_id=school_id).delete()  # cascades entries
            TeachingAssignment.objects.filter(school_class__grade__school_id=school_id).delete()
            Constraint.objects.filter(school_id=school_id).delete()
            summary = {
                'timetables_deleted': tts,
                'entries_deleted': entries,
                'assignments_deleted': asn,
                'constraints_deleted': cons,
            }

    elif op == 'wipe_everything':
        # Full reset to a blank slate for a clean re-import: removes all
        # imported + generated data (timetables, assignments, constraints,
        # roles, tags, teachers, subjects, classes, grades). Keeps the school
        # itself and its structural setup (time slots, rooms) so generation can
        # run again after re-importing.
        from apps.scheduling.models import Constraint
        from apps.school.models import Grade
        from apps.subjects.models import Subject, Teacher, TeacherRole, TeacherTag
        with transaction.atomic():
            summary = {
                'timetables_deleted': Timetable.objects.filter(school_id=school_id).count(),
                'entries_deleted': TimetableEntry.objects.filter(timetable__school_id=school_id).count(),
                'assignments_deleted': TeachingAssignment.objects.filter(
                    school_class__grade__school_id=school_id).count(),
                'constraints_deleted': Constraint.objects.filter(school_id=school_id).count(),
                'roles_deleted': TeacherRole.objects.filter(school_id=school_id).count(),
                'tags_deleted': TeacherTag.objects.filter(school_id=school_id).count(),
                'teachers_deleted': Teacher.objects.filter(school_id=school_id).count(),
                'subjects_deleted': Subject.objects.filter(school_id=school_id).count(),
                'classes_deleted': SchoolClass.objects.filter(grade__school_id=school_id).count(),
                'grades_deleted': Grade.objects.filter(school_id=school_id).count(),
            }
            # Order: referencing rows first.
            Timetable.objects.filter(school_id=school_id).delete()  # cascades entries
            Constraint.objects.filter(school_id=school_id).delete()
            TeachingAssignment.objects.filter(school_class__grade__school_id=school_id).delete()
            TeacherRole.objects.filter(school_id=school_id).delete()
            TeacherTag.objects.filter(school_id=school_id).delete()
            Teacher.objects.filter(school_id=school_id).delete()
            Subject.objects.filter(school_id=school_id).delete()
            SchoolClass.objects.filter(grade__school_id=school_id).delete()
            Grade.objects.filter(school_id=school_id).delete()

    else:
        return Response({'error': f'unknown operation {op!r}'},
                        status=status.HTTP_400_BAD_REQUEST)

    return Response({'ok': True, 'operation': op, 'summary': summary})
