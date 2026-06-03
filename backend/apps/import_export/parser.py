"""
Excel parser for the school's `הערכות לשנת הלימודים` workbook.

The workbook has ~30 sheets. Layouts vary:

  • Most are **subject sheets**: per-class rows with כיתה / סוג כיתה /
    שם המורה המלמד / שעות הוראה / שעות גמול בגרות / סמל שאלון בגרות / הערות.

  • High-school subjects often **pool** multiple classes for ability-level
    groups: a row like ``יא 1,2,5,6,7,9`` followed by several rows for
    ``3 יח"ל / 4 יח"ל / 5 יח"ל``, each with its own teacher and hours. The
    pool rows have a class list in column A; the level rows have a class-
    type label and inherit the most-recent pool.

  • חינוך גופני splits into ``מורה לבנים`` / ``מורה לבנות`` per class —
    we record both as separate assignments under the same group_key.

  • תפקידים lists teacher roles (ניהול, ריכוז, ייעוץ, …) — non-teaching
    hours and stipends, captured into ``TeacherRole``.

  • השכלה is a course catalog (electives) — we capture it loosely; the
    importer doesn't yet bind catalog entries to classes.

The parser is **two-phase**:

  1. ``analyze(file)`` returns a structured ``ParsedWorkbook`` dataclass —
     no database side-effects. Useful for the dry-run preview.

  2. ``apply(parsed, school, log)`` commits the parsed result to the DB,
     transactionally and idempotently keyed on ``source_sheet + source_row``.

The two phases let the FE show a "you're about to import N rows / X
warnings" preview before the user commits.
"""
from __future__ import annotations

import re
import uuid
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Sum
from openpyxl import load_workbook

from apps.school.models import Grade, School, SchoolClass
from apps.subjects.models import Subject, Teacher, TeacherRole, TeachingAssignment


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

GRADE_LETTERS = ['ז', 'ח', 'ט', 'י', 'יא', 'יב']
GRADE_LEVEL = {'ז': 7, 'ח': 8, 'ט': 9, 'י': 10, 'יא': 11, 'יב': 12}
# The same grade letter shows up in various forms.
GRADE_ALIASES = {
    'יו"ד': 'י', 'יוד': 'י', "י'": 'י',
    'י"א': 'יא', "יא'": 'יא',
    'י"ב': 'יב', "יב'": 'יב',
}

# Map raw track labels to the canonical ClassType enum value.
CLASS_TYPE_MAP = {
    'חינוך מיוחד': SchoolClass.ClassType.SPECIAL_ED,
    'מנהיגות': SchoolClass.ClassType.LEADERSHIP,
    ' מנהיגות': SchoolClass.ClassType.LEADERSHIP,
    'מנהיגות ': SchoolClass.ClassType.LEADERSHIP,
    'מופת מדעית': SchoolClass.ClassType.MOFET_SCIENCE,
    'מופת מדעית ': SchoolClass.ClassType.MOFET_SCIENCE,
    'מופ"ת מדעית': SchoolClass.ClassType.MOFET_SCIENCE,
    'מופת מנהיגות': SchoolClass.ClassType.MOFET_LEADERSHIP,
    'מופ"ת מנהיגות': SchoolClass.ClassType.MOFET_LEADERSHIP,
    'עתודה מדעית': SchoolClass.ClassType.RESERVE_SCIENCE,
    'אומ"ץ': SchoolClass.ClassType.OMETS,
    'אומץ': SchoolClass.ClassType.OMETS,
    'מב"ר': SchoolClass.ClassType.MABAR,
    'מבר': SchoolClass.ClassType.MABAR,
    'תלמ': SchoolClass.ClassType.TALM,
    'תל"מ': SchoolClass.ClassType.TALM,
    'אתגר': SchoolClass.ClassType.ATGAR,
    'ביוטכנולוגיה': SchoolClass.ClassType.BIOTECH,
}

ROLE_SHEETS = {'תפקידים'}
COURSE_CATALOG_SHEETS = {'השכלה'}  # captured but not bound to classes
PE_SHEETS = {'חינוך גופני'}
SKIP_SHEETS: set[str] = set()  # explicit skip list — currently nothing

# Sheets that represent **bagrut electives** — courses students opt into
# at the high-school level. Their per-class rows can't be scheduled as
# class-bound lessons (because the audience is a self-selected subset),
# so we import them as inactive and let the school book them into a
# designated "bagrut window" manually. Core subjects are scheduled
# normally.
ELECTIVE_SHEETS = {
    'מדעי החברה', 'ביוטכנולוגיה', 'תעשייה וניהול', 'מנהל וכלכלה',
    'ביולוגיה', 'כימיה', 'פיזיקה', 'מדעי המחשב', 'אדריכלות',
    'פיקוד ובקרה', 'מחשבת ישראל',
}
# Only relevant for high-school grades.
HIGH_SCHOOL_GRADES = {'י', 'יא', 'יב'}

# Headers we recognize. Listed most-specific first so that "סוג כיתה"
# doesn't get matched as "כיתה" first.
HEADER_KEYS: list[tuple[str, list[str]]] = [
    ('class_type', ['סוג כיתה']),
    ('boys_teacher', ['מורה לבנים']),
    ('girls_teacher', ['מורה לבנות']),
    ('boys_count', ['מספרי בנים', 'מספר בנים']),
    ('girls_count', ['מספרי בנות', 'מספר בנות']),
    ('student_count', ["מס' תלמידים", 'מספר תלמידים']),
    ('teacher', ['שם המורה המלמד', 'שם מורה', 'שם המורה', 'מורה']),
    ('bagrut_hours', ['שעות גמול בגרות', 'גמול בגרות']),
    ('bagrut_code', ['סמל שאלון בגרות', 'סמל שאלון']),
    ('hours', ['שעות הוראה', 'שעות']),
    ('notes', ['הערות']),
    ('class', ['כיתה']),
]


# ─────────────────────────────────────────────────────────────────────────────
# Dataclasses
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class AssignmentRow:
    """One delivery row, possibly across multiple classes for a pooled group."""
    source_sheet: str
    source_row: int
    subject: str
    classes: list[tuple[str, int]]  # [(grade_letter, class_number), …]
    class_type_raw: str
    teacher: str
    weekly_hours: Decimal | None
    bagrut_hours: Decimal | None
    bagrut_code: str
    track_label: str
    notes: str
    student_count: int | None
    group_key: str
    is_active: bool
    # PE-specific: a row can describe both a boys teacher and a girls teacher,
    # which we want to import as two assignments. We materialize that into two
    # AssignmentRow records before committing — keeping the dataclass simple.


@dataclass
class RoleRow:
    source_sheet: str
    source_row: int
    role_title: str
    context: str
    description: str
    teacher: str
    weekly_hours: Decimal | None
    stipend_fraction: Decimal | None
    must_teach_hours: Decimal | None
    notes: str


@dataclass
class ParsedWorkbook:
    file_name: str = ''
    assignment_rows: list[AssignmentRow] = field(default_factory=list)
    role_rows: list[RoleRow] = field(default_factory=list)
    sheets_seen: list[str] = field(default_factory=list)
    sheets_unparsed: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
# Cell helpers
# ─────────────────────────────────────────────────────────────────────────────

CLASS_TOKEN_RE = re.compile(r'(יב|יא|י|ט|ח|ז)\s*[\'"]?\s*(\d+)')
CLASS_LIST_PREFIX_RE = re.compile(r'(יב|יא|י|ט|ח|ז)\s*[\'"]?\s*([\d,\s/-]+)')


def _normalize(s) -> str:
    if s is None:
        return ''
    return str(s).strip().replace('"', '"').replace("'", "'")


def _apply_grade_aliases(text: str) -> str:
    for alias, canonical in GRADE_ALIASES.items():
        text = text.replace(alias, canonical)
    return text


def _parse_single_class(text: str):
    text = _apply_grade_aliases(_normalize(text))
    if not text:
        return None
    m = CLASS_TOKEN_RE.match(text)
    if not m:
        return None
    return m.group(1), int(m.group(2))


def _guess_grade(text: str) -> str | None:
    """Return the grade letter implied by a cell, or None.

    Handles partial / fuzzy entries — e.g., 'מופת- ט6' implies ט,
    'בחירת ביולוגיה יא' implies יא. We just look for any grade-letter
    substring and return the longest match (so 'יא' wins over 'י')."""
    text = _apply_grade_aliases(_normalize(text))
    if not text:
        return None
    matches = []
    for g in sorted(GRADE_LEVEL.keys(), key=lambda x: -len(x)):
        if g in text:
            matches.append(g)
    return matches[0] if matches else None


def _parse_class_list(text: str):
    """Parse a cell with a class list, like ``יא 1,2,5,6,7,9`` or ``ח ,2,3,4,8,9``.

    Returns a list of (grade, number) tuples. Empty list if not parseable.
    """
    text = _apply_grade_aliases(_normalize(text))
    if not text:
        return []
    m = CLASS_LIST_PREFIX_RE.match(text)
    if not m:
        return []
    grade = m.group(1)
    nums = re.findall(r'\d+', m.group(2))
    # Heuristic: a stray "92" in ``ח ,92,3,4,8`` is almost certainly two
    # missing-comma numbers ("9, 2"). Class numbers > 12 don't occur, so
    # split any 2-digit number that's outside 1..12 into its digits.
    result = []
    for raw in nums:
        n = int(raw)
        if 1 <= n <= 12 or n == 0:
            result.append((grade, n))
        else:
            # Split "92" → 9, 2; "123" → 1, 2, 3.
            for digit in raw:
                if digit.isdigit():
                    result.append((grade, int(digit)))
    return result


def _parse_hours(value) -> Decimal | None:
    """Parse ``'3+1'`` → ``Decimal('4')``, ``'2.5'`` → ``Decimal('2.5')``, etc.
    Returns None for unparseable values."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = _normalize(value)
    if not s:
        return None
    matches = re.findall(r'\d+(?:\.\d+)?', s)
    if not matches:
        return None
    try:
        return sum((Decimal(m) for m in matches), Decimal('0'))
    except InvalidOperation:
        return None


def _parse_int(value) -> int | None:
    d = _parse_hours(value)
    if d is None:
        return None
    try:
        return int(d)
    except (ValueError, InvalidOperation):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Header / column detection
# ─────────────────────────────────────────────────────────────────────────────

def _match_header_key(text: str) -> str | None:
    text = _normalize(text)
    if not text:
        return None
    # Exact match wins.
    for key, candidates in HEADER_KEYS:
        if any(c == text for c in candidates):
            return key
    # Substring fallback.
    for key, candidates in HEADER_KEYS:
        if any(c in text for c in candidates):
            return key
    return None


def _find_header_row(ws) -> int | None:
    """Find the row that contains a ``כיתה`` cell — that's the header row."""
    for r in range(1, min(10, ws.max_row + 1)):
        for cell in ws[r]:
            if _normalize(cell.value) == 'כיתה':
                return r
    return None


def _detect_columns(ws, header_row: int) -> dict[str, int]:
    """Map header key → column number (1-based). First-occurrence wins per
    key — so a teacher-summary side-table in cols J-P won't clobber the
    primary assignment columns on the left."""
    mapping: dict[str, int] = {}
    for cell in ws[header_row]:
        key = _match_header_key(_normalize(cell.value))
        if key is None:
            continue
        mapping.setdefault(key, cell.column)
    return mapping


def _detect_subject_name(ws) -> str:
    """Find the subject name in the sheet.

    The subject name is consistently in row 1, near the top-left (cols A-E).
    Skips obvious junk: datetime/number cells, header words, 1-char strings.
    Falls back to the sheet title if nothing better is found.
    """
    import datetime as _dt
    header_strings = set()
    for _, candidates in HEADER_KEYS:
        header_strings.update(candidates)
    candidates: list[str] = []
    # Row 1 is the canonical title row. Only look at row 2 if row 1 is empty.
    for r in range(1, 3):
        for cell in ws[r]:
            if cell.column > 6:  # only cols A-F
                break
            # Reject datetime cells — they come from formula cells or
            # accidental date entries (e.g., פיקוד ובקרה has one in the header).
            if isinstance(cell.value, (_dt.datetime, _dt.date)):
                continue
            text = _normalize(cell.value)
            if not text:
                continue
            if text == 'כיתה' or text in header_strings:
                continue
            if re.fullmatch(r'\d+(?:\.\d+)?', text):
                continue
            # 1-char strings are usually data-entry noise (e.g., the "וי" cell
            # in front of אדריכלות).
            if len(text) < 2:
                continue
            candidates.append(text)
        if candidates:
            break
    if candidates:
        return max(candidates, key=len)
    return ws.title


# ─────────────────────────────────────────────────────────────────────────────
# Per-sheet parsers
# ─────────────────────────────────────────────────────────────────────────────

def _parse_subject_sheet(ws, parsed: ParsedWorkbook):
    header_row = _find_header_row(ws)
    if not header_row:
        parsed.sheets_unparsed.append(ws.title)
        parsed.warnings.append(f'{ws.title}: לא נמצאה שורת כותרת')
        return

    cols = _detect_columns(ws, header_row)
    subject = _detect_subject_name(ws)
    is_pe = ws.title in PE_SHEETS or ('boys_teacher' in cols and 'girls_teacher' in cols)

    last_pool: list[tuple[str, int]] = []
    last_pool_key = ''  # group_key shared across rows of one pool

    for r in range(header_row + 1, ws.max_row + 1):
        row = ws[r]
        if not any(_normalize(c.value) for c in row):
            continue

        def cv(key: str):
            col = cols.get(key)
            return row[col - 1].value if col else None

        class_raw = _normalize(cv('class'))
        class_type_raw = _normalize(cv('class_type'))
        teacher = _normalize(cv('teacher'))
        boys_teacher = _normalize(cv('boys_teacher')) if is_pe else ''
        girls_teacher = _normalize(cv('girls_teacher')) if is_pe else ''
        notes = _normalize(cv('notes'))
        bagrut_code = _normalize(cv('bagrut_code'))

        hours = _parse_hours(cv('hours'))
        bagrut_hours = _parse_hours(cv('bagrut_hours'))
        student_count = _parse_int(cv('student_count'))

        # Skip totals rows like "חטע 194 29.75".
        if class_raw.replace(' ', '') in {'חטע', 'חטב', 'חט"ע', 'חט"ב'}:
            continue

        # Determine which classes this row applies to. A cell with a comma
        # after the grade prefix is a *pool* — list-parse it. Otherwise we
        # try the single-class form.
        classes: list[tuple[str, int]] = []
        track_label = class_type_raw  # default; refined below
        new_pool = False
        # Some row variants force is_active=False (e.g., bagrut electives
        # that need manual scheduling). The default below will OR in
        # פתיחה מותנית marker if relevant.
        forced_inactive = False
        if class_raw:
            multi = _parse_class_list(class_raw)
            if multi and len(multi) > 1:
                classes = list(multi)
                last_pool = list(multi)
                new_pool = True
            else:
                single = _parse_single_class(class_raw)
                if single:
                    classes = [single]
                    last_pool = [single]
                    new_pool = True
                else:
                    # Grade-only indicator ("יוד" / "יא" / "יב") with no
                    # class number — typical of bagrut electives where
                    # students from any class can attend. We attach the
                    # row to the FIRST class of that grade as a placeholder
                    # so the teacher's hours are tracked. The school
                    # adjusts these manually post-solve.
                    grade_only = _apply_grade_aliases(class_raw).strip()
                    grade_only = re.sub(r'[\s,\']+$', '', grade_only)
                    if grade_only in GRADE_LEVEL:
                        # Bagrut elective for the whole grade — students
                        # from any class can attend. We import it as
                        # *inactive* so the solver ignores it (the school
                        # books these into a dedicated "bagrut window" by
                        # hand). The teacher's hours still show up via
                        # `TeachingAssignment.weekly_hours`.
                        parsed.warnings.append(
                            f'{ws.title} R{r}: שורת בחירה לשכבה {grade_only} —'
                            f' מסומנת כלא-פעילה (יש לתזמן ידנית)'
                        )
                        classes = [(grade_only, 1)]
                        last_pool = list(classes)
                        new_pool = True
                        forced_inactive = True
                        track_label = (track_label + ' / בחירה' if track_label else 'בחירה')
                    elif last_pool and last_pool[0][0] == _guess_grade(class_raw):
                        # Continuation of a pool whose label is shorthand
                        # for the same grade (e.g., "מופת- ט6" while in ט).
                        notes = (notes + ' | ' if notes else '') + f'(תווית: {class_raw})'
                        classes = list(last_pool)
                    else:
                        # Genuinely unparseable, different grade than the
                        # previous pool — skip and warn.
                        parsed.warnings.append(
                            f'{ws.title} R{r}: לא ניתן לפענח שורת כיתה: {class_raw!r}'
                        )
                        continue
        else:
            # Continuation row — uses the most-recent pool.
            classes = list(last_pool)

        if new_pool:
            last_pool_key = f'{ws.title}#{r}-{uuid.uuid4().hex[:6]}'

        if not classes:
            # Nothing to attach this row to. Skip silently unless it has
            # meaningful content.
            if teacher or hours or class_type_raw:
                parsed.warnings.append(
                    f'{ws.title} R{r}: שורה ללא כיתה ידועה — מדלגים'
                )
            continue

        # PE: split into two assignments (boys / girls) keyed under the same group.
        teachers_to_record: list[tuple[str, str]] = []
        if is_pe:
            if boys_teacher:
                teachers_to_record.append((boys_teacher, 'בנים'))
            if girls_teacher:
                teachers_to_record.append((girls_teacher, 'בנות'))
            # Fall back to plain `teacher` column if neither boys/girls filled.
            if not teachers_to_record and teacher:
                teachers_to_record.append((teacher, ''))
        else:
            teachers_to_record.append((teacher, ''))

        # "פתיחה מותנית" (conditional opening) marks rows where the lesson
        # may or may not actually happen — flag and import as inactive.
        is_active = ('פתיחה מותנית' not in notes) and not forced_inactive
        # Bagrut electives in high school can't be scheduled as class-bound
        # lessons (the audience is self-selected). Mark inactive so the
        # solver ignores them; the school books them in a bagrut window.
        if is_active and ws.title in ELECTIVE_SHEETS and classes:
            grade_letter = classes[0][0]
            if grade_letter in HIGH_SCHOOL_GRADES:
                is_active = False

        # group_key links rows that should be scheduled in parallel:
        #   - a multi-class pool (clearly)
        #   - a single-class block with multiple ability tracks (multiple
        #     continuation rows attached to the same parent row — students
        #     split into level groups during the same time slots)
        # We assign group_key whenever the row is part of a continuation
        # chain — i.e., when last_pool_key was set and either this is a
        # continuation row or there's already more than one row in the
        # chain. The first row of a chain might end up "soloing" with
        # an empty group_key, which is fine; if more continuation rows
        # follow, the parent's group_key is re-used.
        gk = ''
        if last_pool_key:
            # We always set the group_key when last_pool_key exists —
            # this lets the solver treat single-class ability-track
            # families (e.g., י7's 5/4-יח"ל English) as parallel tracks.
            gk = last_pool_key

        for tname, gender_tag in teachers_to_record:
            track = track_label
            if gender_tag:
                track = (track + ' / ' if track else '') + gender_tag
            row_obj = AssignmentRow(
                source_sheet=ws.title,
                source_row=r,
                subject=subject,
                classes=list(classes),
                class_type_raw=class_type_raw,
                teacher=tname,
                weekly_hours=hours,
                bagrut_hours=bagrut_hours,
                bagrut_code=bagrut_code,
                track_label=track,
                notes=notes,
                student_count=student_count,
                group_key=gk,
                is_active=is_active,
            )
            parsed.assignment_rows.append(row_obj)


def _parse_role_sheet(ws, parsed: ParsedWorkbook):
    # Header is row 1. Columns: 1=role, 2=context, 3=description, 4=teacher,
    # 5=weekly hours, 6=stipend, 7=must-teach.
    for r in range(2, ws.max_row + 1):
        cells = [_normalize(c.value) for c in ws[r]]
        if not any(cells):
            continue
        role = cells[0] if len(cells) > 0 else ''
        teacher = cells[3] if len(cells) > 3 else ''
        if not role and not teacher:
            continue
        parsed.role_rows.append(RoleRow(
            source_sheet=ws.title,
            source_row=r,
            role_title=role,
            context=cells[1] if len(cells) > 1 else '',
            description=cells[2] if len(cells) > 2 else '',
            teacher=teacher,
            weekly_hours=_parse_hours(cells[4]) if len(cells) > 4 else None,
            stipend_fraction=_parse_hours(cells[5]) if len(cells) > 5 else None,
            must_teach_hours=_parse_hours(cells[6]) if len(cells) > 6 else None,
            notes=cells[7] if len(cells) > 7 else '',
        ))


def analyze(file, file_name: str = '') -> ParsedWorkbook:
    """First phase — parse the workbook into a ParsedWorkbook without
    touching the DB. Safe to call for a dry-run preview."""
    wb = load_workbook(file, data_only=True)
    parsed = ParsedWorkbook(file_name=file_name)

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        parsed.sheets_seen.append(sheet_name)
        try:
            if sheet_name in ROLE_SHEETS:
                _parse_role_sheet(ws, parsed)
            elif sheet_name in COURSE_CATALOG_SHEETS:
                # We don't yet bind catalog entries to classes — just note it.
                parsed.warnings.append(
                    f'{sheet_name}: גיליון קורסי השכלה — לא מיובא כרגע (יבוא ידני)'
                )
            else:
                _parse_subject_sheet(ws, parsed)
        except Exception as exc:  # parser bug or wildly malformed sheet
            parsed.errors.append(f'{sheet_name}: שגיאת ניתוח: {exc}')

    return parsed


# ─────────────────────────────────────────────────────────────────────────────
# DB application (commit phase)
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_grades(school: School) -> dict[str, Grade]:
    grades: dict[str, Grade] = {}
    for letter, level in GRADE_LEVEL.items():
        g, _ = Grade.objects.get_or_create(
            school=school, level=level,
            defaults={'name': letter, 'order': level},
        )
        if g.name != letter:
            g.name = letter
            g.save(update_fields=['name'])
        grades[letter] = g
    return grades


def _canonical_class_type(raw: str) -> str:
    if not raw:
        return SchoolClass.ClassType.REGULAR
    key = raw.strip()
    if key in CLASS_TYPE_MAP:
        return CLASS_TYPE_MAP[key]
    # Try whitespace-normalized version.
    norm = re.sub(r'\s+', ' ', key)
    if norm in CLASS_TYPE_MAP:
        return CLASS_TYPE_MAP[norm]
    return SchoolClass.ClassType.OTHER


def _canonicalize_teacher_name(raw_name: str) -> str:
    """Strip inline annotations from a teacher cell so the same teacher
    isn't stored multiple times. Common patterns we see in the workbook:

        "אולגה עם ח7"      → "אולגה"
        "אולגה- ח,ט"        → "אולגה"
        "אולגה-יב3"         → "אולגה"
        "סוזנה- יא5"        → "סוזנה"
        "מאיה ק- יא3"       → "מאיה ק"
        "?פאני"             → "פאני"
        "אגי-"              → "אגי"

    The annotations are useful metadata (which class is paired with whom)
    but for the *canonical name* we want just the human-readable name.
    """
    s = raw_name.strip()
    if not s:
        return ''
    # Trim trailing "?" used to mark uncertainty (e.g., "?פאני").
    s = s.strip('?').strip()
    # Strip parenthetical annotations like "(*2)" — they're allocation
    # multipliers, not part of the name. (We don't import the multiplier
    # info; the planning team manages it manually.)
    s = re.sub(r'\s*\([^)]*\)\s*', ' ', s).strip()
    # Cut off everything after a dash followed by a class indicator —
    # e.g., "אולגה-יב3", "סוזנה- יא5", "אגי-" → keep prefix only.
    cut = re.search(r'\s*-\s*', s)
    if cut:
        s = s[:cut.start()].strip()
    # Cut off " עם <anything>" (e.g., "אולגה עם ח7").
    cut = re.search(r'\s+עם\s+', s)
    if cut:
        s = s[:cut.start()].strip()
    # Collapse internal whitespace.
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _valid_canonical(raw_name: str) -> str | None:
    """Canonicalize and reject non-name junk (empty, single non-alpha char,
    pure numbers). Returns the canonical name or None."""
    canonical = _canonicalize_teacher_name(raw_name)
    if not canonical:
        return None
    if len(canonical) < 2 and not canonical.isalpha():
        return None
    if re.fullmatch(r'\d+(?:\.\d+)?', canonical):
        return None
    return canonical


_HEB_LETTER = '[א-ת]'


def _split_teacher_name(canonical: str) -> tuple[str, str]:
    """Split a canonical name into (first, last). A trailing single Hebrew
    letter is a disambiguator that stays part of the first name
    ("מאיה ק" → ("מאיה ק", "")), not a surname."""
    tokens = canonical.split()
    if len(tokens) <= 1:
        return canonical, ''
    if re.fullmatch(_HEB_LETTER + "['׳]?", tokens[-1]):
        return canonical, ''
    return tokens[0], ' '.join(tokens[1:])


def _full_name(teacher: Teacher) -> str:
    return f'{teacher.first_name} {teacher.last_name}'.strip()


def _build_existing_index(school: School) -> dict[str, list[Teacher]]:
    """Index existing teachers by the first token of their name (splitting
    legacy full-name-in-first_name rows the same way) so an incoming
    first-name-only cell can find them."""
    index: dict[str, list[Teacher]] = {}
    for t in Teacher.objects.filter(school=school).order_by('id'):
        key = _split_teacher_name(_full_name(t))[0]
        index.setdefault(key, []).append(t)
    return index


def resolve_teacher_match(canonical: str, existing_index: dict[str, list[Teacher]]) -> dict:
    """Decide whether a canonical incoming name matches an existing teacher,
    is new, or is ambiguous. Returns {status, canonical, first, last,
    candidates:[{id, display_name}], suggested} where suggested is a teacher
    id or 'new'. Shared by the dry-run preview and the commit path."""
    first, last = _split_teacher_name(canonical)
    candidates = existing_index.get(first, [])
    cand_list = [{'id': c.id, 'display_name': str(c)} for c in candidates]

    # Exact full-name match (same first token + same last name).
    for c in candidates:
        cf, cl = _split_teacher_name(_full_name(c))
        if cf == first and cl == last:
            return {'status': 'matched', 'canonical': canonical, 'first': first,
                    'last': last, 'candidates': cand_list, 'suggested': c.id}

    if not candidates:
        return {'status': 'new', 'canonical': canonical, 'first': first,
                'last': last, 'candidates': [], 'suggested': 'new'}

    if len(candidates) == 1:
        only = candidates[0]
        _, ol = _split_teacher_name(_full_name(only))
        if not last and not ol:
            # both bare first names, identical → same person
            return {'status': 'matched', 'canonical': canonical, 'first': first,
                    'last': last, 'candidates': cand_list, 'suggested': only.id}
        # one side carries a last name, the other doesn't → likely same, ask
        return {'status': 'ambiguous', 'canonical': canonical, 'first': first,
                'last': last, 'candidates': cand_list, 'suggested': only.id}

    # several existing teachers share the first token → never silently pick one
    return {'status': 'ambiguous', 'canonical': canonical, 'first': first,
            'last': last, 'candidates': cand_list, 'suggested': 'new'}


def _teacher_for(
    school: School,
    raw_name: str,
    cache: dict[str, Teacher],
    *,
    subject_key: str = '',  # kept for call-site compatibility; ignored
    overrides: dict | None = None,
    existing_index: dict[str, list[Teacher]] | None = None,
    warnings: list[str] | None = None,
) -> Teacher | None:
    """Resolve a raw teacher cell to a Teacher, matching on first + last name.

    Order: an explicit user override (keyed by canonical name) wins; otherwise
    ``resolve_teacher_match`` decides match / new / ambiguous and we apply its
    suggested default (merge a single candidate, else create new). Ambiguous
    names are surfaced in the import review step so the user can override before
    commit. New full-name cells store a split first/last; merging a full name
    into a bare first-name record backfills the empty last name (never
    overwrites, never touches legacy full-name-in-first_name rows).
    """
    canonical = _valid_canonical(raw_name)
    if not canonical:
        return None

    cached = cache.get(canonical)
    if cached:
        return cached

    if existing_index is None:
        existing_index = _build_existing_index(school)
    first, last = _split_teacher_name(canonical)

    # 1) explicit user override: teacher id, 'new', or an alias to another
    #    incoming canonical name (for merging two names that are both new).
    if overrides and canonical in overrides:
        val = overrides[canonical]
        if val == 'new':
            t = Teacher.objects.create(school=school, first_name=first, last_name=last)
            cache[canonical] = t
            return t
        if isinstance(val, int) or (isinstance(val, str) and val.isdigit()):
            t = Teacher.objects.filter(id=int(val), school=school).first()
            if t:
                cache[canonical] = t
                return t
            if warnings is not None:
                warnings.append(f'בחירת מורה לא תקפה עבור "{canonical}" — שובץ אוטומטית')
        elif isinstance(val, str) and val and val != canonical:
            alias = cache.get(val) or _teacher_for(
                school, val, cache, overrides=overrides,
                existing_index=existing_index, warnings=warnings,
            )
            if alias:
                cache[canonical] = alias
                return alias
        # bad override → fall through to auto resolution

    # 2) auto resolution
    res = resolve_teacher_match(canonical, existing_index)
    target = res['suggested']
    if target != 'new':
        t = Teacher.objects.filter(id=target, school=school).first()
        if t:
            if last and not t.last_name and t.first_name == first:
                t.last_name = last
                t.save(update_fields=['last_name'])
            cache[canonical] = t
            return t

    t = Teacher.objects.create(school=school, first_name=first, last_name=last)
    cache[canonical] = t
    return t


@dataclass
class ApplyResult:
    subjects_created: int = 0
    teachers_created: int = 0
    classes_created: int = 0
    assignments_created: int = 0
    assignments_updated: int = 0
    roles_created: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def apply(
    parsed: ParsedWorkbook, school: School, *,
    wipe_existing: bool = False, teacher_overrides: dict | None = None,
) -> ApplyResult:
    """Commit a ParsedWorkbook to the database. Returns a structured summary
    of what was inserted/updated.

    Idempotent: assignments are upserted keyed on
    (school, source_sheet, source_row, school_class, teacher, subject)
    so re-uploading the same Excel does not duplicate rows.

    When ``wipe_existing`` is True, all teacher/subject/assignment/role
    records for the school are deleted before re-importing — useful when
    the user wants to fully replace prior data.

    ``teacher_overrides`` maps a canonical incoming teacher name to the
    user's identity decision from the import review step: an existing
    Teacher id, the string ``'new'``, or another incoming canonical name to
    merge with. Absent → automatic first+last matching."""

    result = ApplyResult()
    grades = _ensure_grades(school)

    teacher_cache: dict[str, Teacher] = {}
    subject_cache: dict[str, Subject] = {}
    seen_classes: set[tuple[str, int]] = set()
    seen_teachers_before = set(Teacher.objects.filter(school=school).values_list('id', flat=True))
    seen_subjects_before = set(Subject.objects.filter(school=school).values_list('id', flat=True))

    with transaction.atomic():
        if wipe_existing:
            # Order matters — clear referencing rows first.
            TeachingAssignment.objects.filter(subject__school=school).delete()
            TeacherRole.objects.filter(school=school).delete()
            # We deliberately *keep* SchoolClass rows even on wipe, so that
            # any pre-existing per-class settings (room, homeroom_teacher)
            # survive. Subjects and Teachers re-populate from the Excel.
            Subject.objects.filter(school=school).delete()
            Teacher.objects.filter(school=school).delete()

        # Index existing teachers once (post-wipe) for first+last matching.
        existing_index = _build_existing_index(school)

        # First pass: resolve every (grade, number) referenced into a SchoolClass.
        # We collect them and bulk-fetch to avoid N queries per row.
        for row in parsed.assignment_rows:
            for g, n in row.classes:
                seen_classes.add((g, n))

        # Create missing classes.
        for (g_letter, num) in sorted(seen_classes):
            grade = grades.get(g_letter)
            if not grade:
                result.warnings.append(f'שכבה לא ידועה: {g_letter}')
                continue
            cls, created = SchoolClass.objects.get_or_create(
                grade=grade, number=num,
            )
            if created:
                result.classes_created += 1

        # Build a cache (grade_letter, number) → SchoolClass for fast lookup.
        cls_cache: dict[tuple[str, int], SchoolClass] = {}
        for cls in SchoolClass.objects.filter(grade__school=school).select_related('grade'):
            cls_cache[(cls.grade.name, cls.number)] = cls

        # Second pass: create assignments.
        # Pool rows (len(classes) > 1) create ONE TeachingAssignment attached
        # to the first class as the primary, with the rest in additional_classes.
        # This keeps total teacher hours correct (one delivery = one record).
        for row in parsed.assignment_rows:
            if row.weekly_hours is None or row.weekly_hours == 0:
                # No hours specified → skip but warn (likely placeholder cell).
                if row.teacher or row.bagrut_code:
                    result.warnings.append(
                        f'{row.source_sheet} R{row.source_row}: שורה ללא שעות הוראה'
                    )
                continue

            # Subject lookup (cached per workbook).
            subject = subject_cache.get(row.subject)
            if not subject:
                subject, _ = Subject.objects.get_or_create(
                    school=school, name_he=row.subject,
                )
                subject_cache[row.subject] = subject

            teacher = _teacher_for(
                school, row.teacher, teacher_cache, subject_key=row.subject,
                overrides=teacher_overrides, existing_index=existing_index,
                warnings=result.warnings,
            )

            # Resolve all classes for this row.
            resolved_classes = []
            for (g_letter, num) in row.classes:
                cls = cls_cache.get((g_letter, num))
                if not cls:
                    result.warnings.append(
                        f'{row.source_sheet} R{row.source_row}: כיתה לא נמצאה: {g_letter}{num}'
                    )
                    continue
                resolved_classes.append(cls)

            if not resolved_classes:
                continue

            # Set class type from the row hint, but only for classes that
            # are still in the default REGULAR state — don't clobber an
            # explicit type a prior sheet already set.
            for cls in resolved_classes:
                if row.class_type_raw and cls.class_type == SchoolClass.ClassType.REGULAR:
                    canonical = _canonical_class_type(row.class_type_raw)
                    if canonical != SchoolClass.ClassType.REGULAR:
                        cls.class_type = canonical
                        cls.track_label_raw = row.class_type_raw
                        cls.save(update_fields=['class_type', 'track_label_raw'])

            primary = resolved_classes[0]
            extras = resolved_classes[1:]

            obj, created = TeachingAssignment.objects.update_or_create(
                subject=subject,
                school_class=primary,
                source_sheet=row.source_sheet,
                source_row=row.source_row,
                teacher=teacher,
                defaults={
                    'weekly_hours': row.weekly_hours,
                    'bagrut_bonus_hours': row.bagrut_hours or Decimal('0'),
                    'bagrut_exam_code': row.bagrut_code,
                    'track_label': row.track_label,
                    'group_key': row.group_key,
                    'student_count': row.student_count,
                    'is_active': row.is_active,
                    'notes': row.notes,
                },
            )
            if extras:
                obj.additional_classes.set(extras)
            else:
                obj.additional_classes.clear()
            if created:
                result.assignments_created += 1
            else:
                result.assignments_updated += 1

        # Third pass: homeroom teachers. The חינוך sheet maps each class →
        # its homeroom teacher. Capture that into SchoolClass.homeroom_teacher.
        for row in parsed.assignment_rows:
            if row.source_sheet != 'חינוך':
                continue
            if not row.teacher or len(row.classes) != 1:
                continue
            teacher = _teacher_for(
                school, row.teacher, teacher_cache, subject_key=row.subject,
                overrides=teacher_overrides, existing_index=existing_index,
                warnings=result.warnings,
            )
            if not teacher:
                continue
            g_letter, num = row.classes[0]
            cls = cls_cache.get((g_letter, num))
            if cls and cls.homeroom_teacher_id != teacher.id:
                cls.homeroom_teacher = teacher
                cls.save(update_fields=['homeroom_teacher'])

        # Fourth pass: roles. The role sheet uses full teacher names
        # ("רינת שוורץ" etc.) without subject context, so we pass an
        # empty subject_key — letting the resolver pick the first
        # existing Teacher with that name (creating one if none).
        for rrow in parsed.role_rows:
            teacher = _teacher_for(
                school, rrow.teacher, teacher_cache,
                overrides=teacher_overrides, existing_index=existing_index,
                warnings=result.warnings,
            ) if rrow.teacher else None
            TeacherRole.objects.update_or_create(
                school=school,
                role_title=rrow.role_title,
                context=rrow.context,
                teacher=teacher,
                defaults={
                    'description': rrow.description,
                    'weekly_hours': rrow.weekly_hours or Decimal('0'),
                    'stipend_fraction': rrow.stipend_fraction or Decimal('0'),
                    'must_teach_hours': rrow.must_teach_hours or Decimal('0'),
                    'notes': rrow.notes,
                },
            )
            result.roles_created += 1

        # Detect newly-created teachers / subjects (by diffing IDs).
        result.teachers_created = (
            Teacher.objects.filter(school=school).exclude(id__in=seen_teachers_before).count()
        )
        result.subjects_created = (
            Subject.objects.filter(school=school).exclude(id__in=seen_subjects_before).count()
        )

        # Refresh each teacher's max_weekly_hours to the sum of their active
        # assigned hours (rounded up to int — model field is integer for now).
        for teacher in Teacher.objects.filter(school=school):
            total = (
                teacher.assignments.filter(is_active=True).aggregate(s=Sum('weekly_hours'))['s']
                or Decimal('0')
            )
            new_max = int(total) + (1 if total - int(total) > 0 else 0)
            if new_max != teacher.max_weekly_hours:
                teacher.max_weekly_hours = max(new_max, 1)
                teacher.save(update_fields=['max_weekly_hours'])

        # Auto-categorize: department tag per primary subject + coordinator
        # tags from the roles sheet. Idempotent and re-runnable; fills in
        # departments only where empty so manual moves survive a re-import.
        from apps.subjects.services.tagging import sync_all_tags
        sync_all_tags(school)

    result.warnings.extend(parsed.warnings)
    result.errors.extend(parsed.errors)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Legacy entry point (used by the existing views.py)
# ─────────────────────────────────────────────────────────────────────────────

def parse_timetable_excel(file, school_id, import_log):
    """Backwards-compatible thin shim — analyzes the workbook and immediately
    commits. Updates ``import_log`` with totals and warnings.

    Prefer the two-phase ``analyze`` / ``apply`` API directly for new code.
    """
    parsed = analyze(file, file_name=getattr(file, 'name', ''))
    school = School.objects.get(id=school_id)
    result = apply(parsed, school)

    import_log.subjects_imported = result.subjects_created
    import_log.teachers_imported = result.teachers_created
    import_log.classes_imported = result.classes_created
    import_log.assignments_imported = result.assignments_created + result.assignments_updated
    import_log.roles_imported = result.roles_created
    import_log.errors = result.errors
    import_log.warnings = result.warnings
    import_log.details = {
        'sheets_seen': parsed.sheets_seen,
        'assignment_rows_parsed': len(parsed.assignment_rows),
        'role_rows_parsed': len(parsed.role_rows),
        'assignments_created': result.assignments_created,
        'assignments_updated': result.assignments_updated,
    }
    import_log.save()

    return {
        'subjects': result.subjects_created,
        'teachers': result.teachers_created,
        'assignments': result.assignments_created + result.assignments_updated,
        'errors': result.errors,
        'warnings': result.warnings,
    }
